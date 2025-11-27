import openpyxl

def analyze_excel_structure():
    try:
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
        final_wb = openpyxl.load_workbook("Excel_Data/Exams/Final Term.xlsx")
        
        class_sheets = ['Class  II', 'Class III', 'Class IV', 'Class V', 'Class VI', 'Class VII', 'Class VIII', 'Class IX', 'Class X']
        
        for sheet_name in class_sheets:
            if sheet_name in mid_wb.sheetnames:
                print(f"\n{'='*60}")
                print(f"ANALYZING {sheet_name}")
                print(f"{'='*60}")
                
                mid_sheet = mid_wb[sheet_name]
                
                # Find header row by looking for subject names
                header_row = None
                for row in range(1, 10):
                    for col in range(1, 15):
                        cell_val = mid_sheet.cell(row=row, column=col).value
                        if cell_val and any(subject in str(cell_val).upper() for subject in ['ENGLISH', 'URDU', 'MATH', 'SCIENCE', 'COMPUTER']):
                            header_row = row
                            break
                    if header_row:
                        break
                
                if header_row:
                    print(f"Header row found at: {header_row}")
                    print("Column headers:")
                    for col in range(1, 15):
                        cell_val = mid_sheet.cell(row=header_row, column=col).value
                        if cell_val:
                            col_letter = openpyxl.utils.get_column_letter(col)
                            print(f"  {col_letter}: {cell_val}")
                else:
                    print("No clear header row found. Checking first few students:")
                
                # Find first student and analyze their data
                student_found = False
                for row in range(1, 20):
                    for col in ['A', 'B', 'C']:
                        cell_val = mid_sheet[f'{col}{row}'].value
                        if cell_val and str(cell_val).startswith('AAH'):
                            next_col = chr(ord(col) + 1)
                            name_val = mid_sheet[f'{next_col}{row}'].value
                            if name_val and not str(name_val).startswith('AAH'):
                                print(f"\nFirst student: {cell_val} - {name_val}")
                                print(f"Student data at row {row}:")
                                
                                # Show all data columns
                                for mark_col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                                    mark_val = mid_sheet[f'{mark_col}{row}'].value
                                    if mark_val is not None:
                                        print(f"  {mark_col}: {mark_val} ({type(mark_val).__name__})")
                                
                                student_found = True
                                break
                    if student_found:
                        break
                
                if not student_found:
                    print("No student data found in this sheet")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    analyze_excel_structure()
