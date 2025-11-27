#!/usr/bin/env python3
"""
Excel Sheet Access Verification Tool
===================================
Verifies that data is being fetched from correct class sheets in both Excel files
"""

import openpyxl
import json

def verify_excel_access():
    print("🔍 EXCEL SHEET ACCESS VERIFICATION")
    print("=" * 60)
    
    try:
        # Load both Excel files
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
        final_wb = openpyxl.load_workbook("Excel_Data/Exams/Final Term.xlsx")
        
        print("✅ Both Excel files loaded successfully")
        print(f"📁 Mid Term sheets: {mid_wb.sheetnames}")
        print(f"📁 Final Term sheets: {final_wb.sheetnames}")
        print()
        
        # Expected class sheets
        expected_sheets = ['Class  II', 'Class III', 'Class IV', 'Class V', 'Class VI', 'Class VII', 'Class VIII', 'Class IX', 'Class X']
        
        verification_log = {}
        
        for sheet_name in expected_sheets:
            print(f"🔍 Verifying {sheet_name}...")
            
            sheet_data = {
                'sheet_name': sheet_name,
                'mid_exists': sheet_name in mid_wb.sheetnames,
                'final_exists': sheet_name in final_wb.sheetnames,
                'students_found': [],
                'sample_data': {}
            }
            
            if sheet_name in mid_wb.sheetnames and sheet_name in final_wb.sheetnames:
                mid_sheet = mid_wb[sheet_name]
                final_sheet = final_wb[sheet_name]
                
                print(f"  ✅ Sheet exists in both files")
                
                # Check for students in this sheet
                students_in_sheet = []
                
                for row in range(1, 50):  # Check first 50 rows
                    for col in ['A', 'B', 'C']:
                        try:
                            cell_val = mid_sheet[f'{col}{row}'].value
                            if cell_val and str(cell_val).startswith('AAH'):
                                next_col = chr(ord(col) + 1)
                                name_val = mid_sheet[f'{next_col}{row}'].value
                                if name_val and not str(name_val).startswith('AAH'):
                                    student_id = str(cell_val).strip()
                                    student_name = str(name_val).strip()
                                    
                                    # Get sample data from both sheets
                                    mid_marks = []
                                    final_marks = []
                                    
                                    for mark_col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
                                        mid_val = mid_sheet[f'{mark_col}{row}'].value
                                        final_val = final_sheet[f'{mark_col}{row}'].value
                                        mid_marks.append(mid_val)
                                        final_marks.append(final_val)
                                    
                                    student_data = {
                                        'id': student_id,
                                        'name': student_name,
                                        'row': row,
                                        'col': col,
                                        'mid_marks': mid_marks,
                                        'final_marks': final_marks,
                                        'class': sheet_name.replace('Class ', '').strip()
                                    }
                                    
                                    students_in_sheet.append(student_data)
                                    
                                    # Store first 3 students as sample
                                    if len(sheet_data['sample_data']) < 3:
                                        sheet_data['sample_data'][student_id] = student_data
                                    
                                    break
                        except:
                            continue
                
                sheet_data['students_found'] = len(students_in_sheet)
                print(f"  👥 Found {len(students_in_sheet)} students")
                
                # Show sample data
                if sheet_data['sample_data']:
                    print(f"  📊 Sample students:")
                    for student_id, data in list(sheet_data['sample_data'].items())[:2]:
                        print(f"    {student_id} - {data['name']} (Row {data['row']})")
                        print(f"      Mid marks: {data['mid_marks'][:5]}...")
                        print(f"      Final marks: {data['final_marks'][:5]}...")
                
            else:
                print(f"  ❌ Sheet missing - Mid: {sheet_name in mid_wb.sheetnames}, Final: {sheet_name in final_wb.sheetnames}")
            
            verification_log[sheet_name] = sheet_data
            print()
        
        # Save verification log
        with open('excel_access_verification.json', 'w') as f:
            json.dump(verification_log, f, indent=2)
        
        # Summary
        print("📋 VERIFICATION SUMMARY:")
        print("-" * 40)
        total_students = 0
        for sheet_name, data in verification_log.items():
            status = "✅" if data['mid_exists'] and data['final_exists'] else "❌"
            print(f"{status} {sheet_name}: {data['students_found']} students")
            total_students += data['students_found']
        
        print(f"\n🎯 Total students found: {total_students}")
        print(f"📁 Verification log saved: excel_access_verification.json")
        
        return verification_log
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return None

def verify_specific_student(student_id):
    """Verify data access for a specific student"""
    print(f"🔍 VERIFYING SPECIFIC STUDENT: {student_id}")
    print("=" * 50)
    
    try:
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
        final_wb = openpyxl.load_workbook("Excel_Data/Exams/Final Term.xlsx")
        
        class_sheets = ['Class  II', 'Class III', 'Class IV', 'Class V', 'Class VI', 'Class VII', 'Class VIII', 'Class IX', 'Class X']
        
        for sheet_name in class_sheets:
            if sheet_name in mid_wb.sheetnames and sheet_name in final_wb.sheetnames:
                mid_sheet = mid_wb[sheet_name]
                final_sheet = final_wb[sheet_name]
                
                for row in range(1, 200):
                    for col in ['A', 'B', 'C']:
                        try:
                            cell_val = mid_sheet[f'{col}{row}'].value
                            if cell_val and str(cell_val).strip() == student_id:
                                next_col = chr(ord(col) + 1)
                                name_val = mid_sheet[f'{next_col}{row}'].value
                                
                                print(f"✅ FOUND in {sheet_name}")
                                print(f"📍 Location: Row {row}, Column {col}")
                                print(f"👤 Name: {name_val}")
                                print(f"🏫 Class: {sheet_name.replace('Class ', '').strip()}")
                                print()
                                
                                print("📊 RAW DATA COMPARISON:")
                                print("-" * 30)
                                
                                for i, mark_col in enumerate(['D', 'E', 'F', 'G', 'H', 'I', 'J']):
                                    mid_val = mid_sheet[f'{mark_col}{row}'].value
                                    final_val = final_sheet[f'{mark_col}{row}'].value
                                    
                                    print(f"Column {mark_col}: Mid={mid_val}, Final={final_val}")
                                
                                print()
                                print("✅ Data access verification complete!")
                                return True
                        except:
                            continue
        
        print(f"❌ Student {student_id} not found in any sheet")
        return False
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def compare_system_vs_excel():
    """Compare system data extraction with direct Excel access"""
    print("🔍 COMPARING SYSTEM VS DIRECT EXCEL ACCESS")
    print("=" * 50)
    
    # Import the system's data loading function
    import sys
    sys.path.append('Source_Code')
    
    try:
        from report_card_fixed_totals import ReportCardFixedTotals
        import tkinter as tk
        
        # Create a temporary instance to load data
        root = tk.Tk()
        root.withdraw()  # Hide the window
        
        app = ReportCardFixedTotals(root)
        system_students = app.students
        
        print(f"📊 System loaded: {len(system_students)} students")
        
        # Compare with direct Excel access
        verification_log = verify_excel_access()
        
        if verification_log:
            excel_total = sum(data['students_found'] for data in verification_log.values())
            print(f"📊 Direct Excel access: {excel_total} students")
            
            if len(system_students) == excel_total:
                print("✅ Student count matches!")
            else:
                print("⚠️ Student count mismatch!")
            
            # Sample comparison
            print("\n🔍 SAMPLE DATA COMPARISON:")
            print("-" * 30)
            
            sample_ids = list(system_students.keys())[:3]
            for student_id in sample_ids:
                if student_id in system_students:
                    system_data = system_students[student_id]
                    print(f"\n👤 {student_id} - {system_data['name']}")
                    print(f"   System Class: {system_data['class']}")
                    print(f"   System Mid: {system_data['mid_marks'][:5]}...")
                    print(f"   System Final: {system_data['final_marks'][:5]}...")
                    
                    # Verify with direct Excel access
                    verify_specific_student(student_id)
        
        root.destroy()
        
    except Exception as e:
        print(f"❌ Error in comparison: {e}")

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        if sys.argv[1] == "compare":
            compare_system_vs_excel()
        else:
            verify_specific_student(sys.argv[1])
    else:
        print("Excel Sheet Access Verification Tool")
        print("=" * 40)
        print("Usage:")
        print("  python3 verify_excel_access.py                    - Verify all sheets")
        print("  python3 verify_excel_access.py <student_id>       - Verify specific student")
        print("  python3 verify_excel_access.py compare            - Compare system vs Excel")
        print()
        print("Examples:")
        print("  python3 verify_excel_access.py 'AAH- 170'")
        print("  python3 verify_excel_access.py compare")
        print()
        
        verify_excel_access()
