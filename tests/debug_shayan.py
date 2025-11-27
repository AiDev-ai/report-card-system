#!/usr/bin/env python3
import openpyxl

def debug_shayan_data():
    print("Debugging Shayan's Data")
    print("=" * 40)
    
    try:
        # Load Excel files
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
        final_wb = openpyxl.load_workbook("Excel_Data/Exams/Final Term.xlsx")
        
        # Check Class I sheet
        if "Class I" in mid_wb.sheetnames:
            mid_sheet = mid_wb["Class I"]
            final_sheet = final_wb["Class I"]
            
            print("📋 Scanning Class I sheet for Shayan...")
            
            # Find Shayan's row
            for row in range(1, 50):
                for col in ['A', 'B', 'C', 'D']:
                    cell_val = mid_sheet[f'{col}{row}'].value
                    if cell_val and 'AAH- 220' in str(cell_val):
                        print(f"✅ Found Shayan at Row {row}, Column {col}")
                        
                        # Check name column
                        for name_offset in [1, 2, 3]:
                            try:
                                next_col = chr(ord(col) + name_offset)
                                name_val = mid_sheet[f'{next_col}{row}'].value
                                if name_val and 'Shayan' in str(name_val):
                                    print(f"✅ Name found in column {next_col}: {name_val}")
                                    
                                    # Check marks columns
                                    print("\n📊 Marks Data:")
                                    print("Col | Mid Term | Final Term")
                                    print("-" * 30)
                                    
                                    start_col_ord = ord(next_col) + 1
                                    for i in range(15):
                                        mark_col = chr(start_col_ord + i)
                                        if ord(mark_col) > ord('Z'):
                                            break
                                            
                                        mid_val = mid_sheet[f'{mark_col}{row}'].value
                                        final_val = final_sheet[f'{mark_col}{row}'].value
                                        
                                        if mid_val is not None or final_val is not None:
                                            print(f" {mark_col}  |    {mid_val}    |     {final_val}")
                                        else:
                                            print(f" {mark_col}  |   None   |    None")
                                            if i > 5:  # Stop after a few empty columns
                                                break
                                    
                                    return
                            except:
                                continue
        else:
            print("❌ Class I sheet not found")
            
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    debug_shayan_data()
