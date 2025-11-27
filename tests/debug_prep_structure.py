#!/usr/bin/env python3
import openpyxl

def debug_prep_structure():
    print("Debugging Prep Class Excel Structure")
    print("=" * 50)
    
    try:
        # Load Excel files
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
        
        # Check Class Prep sheet
        if "Class Prep" in mid_wb.sheetnames:
            mid_sheet = mid_wb["Class Prep"]
            
            print("📋 Class Prep Sheet Structure:")
            print("Row | Col A | Col B | Col C | Col D | Col E | Col F | Col G | Col H | Col I | Col J")
            print("-" * 80)
            
            # Show first 10 rows to understand structure
            for row in range(1, 11):
                row_data = []
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                    cell_val = mid_sheet[f'{col}{row}'].value
                    if cell_val is None:
                        row_data.append("None")
                    else:
                        row_data.append(str(cell_val)[:10])  # Truncate long values
                
                print(f" {row:2} | {' | '.join(f'{val:8}' for val in row_data)}")
            
            print("\n" + "=" * 50)
            print("Looking for student data...")
            
            # Find students
            for row in range(1, 20):
                for col in ['A', 'B', 'C', 'D']:
                    cell_val = mid_sheet[f'{col}{row}'].value
                    if cell_val and 'AAH' in str(cell_val):
                        print(f"Found student at Row {row}, Col {col}: {cell_val}")
                        
                        # Show the entire row
                        print("Full row data:")
                        for check_col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                            val = mid_sheet[f'{check_col}{row}'].value
                            print(f"  {check_col}: {val}")
                        break
                if 'AAH' in str(mid_sheet[f'A{row}'].value or ''):
                    break
        else:
            print("❌ Class Prep sheet not found")
            print("Available sheets:", mid_wb.sheetnames)
            
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    debug_prep_structure()
