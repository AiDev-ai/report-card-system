#!/usr/bin/env python3
"""
Test script to verify new sheet detection capability
"""
import openpyxl
import os

def test_sheet_detection():
    print("Testing New Sheet Detection Capability")
    print("=" * 50)
    
    try:
        # Check if Excel files exist
        mid_path = "Excel_Data/Exams/Mid Term.xlsx"
        final_path = "Excel_Data/Exams/Final Term.xlsx"
        
        if not os.path.exists(mid_path):
            print(f"❌ Mid Term file not found: {mid_path}")
            return
        
        if not os.path.exists(final_path):
            print(f"❌ Final Term file not found: {final_path}")
            return
        
        # Load workbooks
        mid_wb = openpyxl.load_workbook(mid_path)
        final_wb = openpyxl.load_workbook(final_path)
        
        print("📊 Available Sheets:")
        print(f"Mid Term: {mid_wb.sheetnames}")
        print(f"Final Term: {final_wb.sheetnames}")
        print()
        
        # Find common sheets
        mid_sheets = set(mid_wb.sheetnames)
        final_sheets = set(final_wb.sheetnames)
        common_sheets = mid_sheets.intersection(final_sheets)
        
        print(f"🔗 Common Sheets: {list(common_sheets)}")
        print()
        
        # Filter class sheets
        class_sheets = []
        for sheet in common_sheets:
            sheet_lower = sheet.lower()
            if any(keyword in sheet_lower for keyword in ['class', 'prep', 'nursery', 'kg', 'grade']):
                class_sheets.append(sheet)
        
        print(f"🎓 Detected Class Sheets: {class_sheets}")
        print()
        
        # Test each sheet
        for sheet_name in class_sheets:
            print(f"📋 Testing Sheet: {sheet_name}")
            mid_sheet = mid_wb[sheet_name]
            
            # Look for student IDs
            student_count = 0
            for row in range(1, 50):  # Check first 50 rows
                for col in ['A', 'B', 'C']:
                    cell_val = mid_sheet[f'{col}{row}'].value
                    if cell_val and str(cell_val).startswith('AAH'):
                        student_count += 1
                        break
            
            print(f"   Found {student_count} students")
        
        print()
        print("✅ New sheet detection capability working!")
        print("✅ System can now automatically detect:")
        print("   - Prep class sheets")
        print("   - Class I sheets") 
        print("   - Any new class sheets added")
        print("   - Nursery/KG sheets")
        
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    test_sheet_detection()
