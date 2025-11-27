import streamlit as st
import pandas as pd
import openpyxl
from PIL import Image
import base64
import os

st.set_page_config(
    page_title="Report Card System",
    page_icon="🏫",
    layout="wide"
)

# Load CSS for styling
st.markdown("""
<style>
.main-header {
    text-align: center;
    color: #2E86AB;
    font-size: 2.5rem;
    margin-bottom: 2rem;
}
.student-card {
    background: #f0f2f6;
    padding: 1rem;
    border-radius: 10px;
    margin: 1rem 0;
}
</style>
""", unsafe_allow_html=True)

@st.cache_data
def load_students():
    """Load student data from Excel files"""
    try:
        students = {}
        
        # Load Mid Term data
        mid_wb = openpyxl.load_workbook("Mid Term.xlsx")
        final_wb = openpyxl.load_workbook("Final Term.xlsx")
        
        # Get common sheets
        common_sheets = set(mid_wb.sheetnames) & set(final_wb.sheetnames)
        
        for sheet_name in common_sheets:
            mid_sheet = mid_wb[sheet_name]
            final_sheet = final_wb[sheet_name]
            
            # Find header row
            header_row = None
            for row in range(1, 10):
                cell_value = str(mid_sheet.cell(row, 1).value or "").strip()
                if cell_value.startswith("AAH"):
                    header_row = row - 1
                    break
            
            if header_row:
                # Get subjects from header
                subjects = []
                for col in range(3, 15):
                    subject = mid_sheet.cell(header_row, col).value
                    if subject:
                        subjects.append(str(subject).strip())
                
                # Process students
                for row in range(header_row + 1, mid_sheet.max_row + 1):
                    student_id = mid_sheet.cell(row, 1).value
                    student_name = mid_sheet.cell(row, 2).value
                    
                    if student_id and student_name:
                        student_id = str(student_id).strip()
                        student_name = str(student_name).strip()
                        
                        # Get marks
                        mid_marks = []
                        final_marks = []
                        
                        for col in range(3, 3 + len(subjects)):
                            mid_mark = mid_sheet.cell(row, col).value
                            final_mark = final_sheet.cell(row, col).value
                            
                            mid_marks.append(mid_mark if mid_mark is not None else 0)
                            final_marks.append(final_mark if final_mark is not None else 0)
                        
                        students[student_id] = {
                            'name': student_name,
                            'class': sheet_name,
                            'subjects': subjects,
                            'mid_marks': mid_marks,
                            'final_marks': final_marks
                        }
        
        return students
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return {}

def generate_report_card(student_data):
    """Generate HTML report card"""
    html = f"""
    <div style="max-width: 800px; margin: 0 auto; padding: 20px; font-family: Arial;">
        <div style="text-align: center; border: 2px solid #333; padding: 20px;">
            <h1 style="color: #2E86AB;">ALKHIDMAT AGHOSH HALA</h1>
            <h2>REPORT CARD</h2>
            <hr>
            <div style="text-align: left; margin: 20px 0;">
                <p><strong>Student Name:</strong> {student_data['name']}</p>
                <p><strong>Class:</strong> {student_data['class']}</p>
                <p><strong>Session:</strong> 2024-25</p>
            </div>
            
            <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                <tr style="background: #2E86AB; color: white;">
                    <th style="border: 1px solid #333; padding: 8px;">Subject</th>
                    <th style="border: 1px solid #333; padding: 8px;">Mid Term</th>
                    <th style="border: 1px solid #333; padding: 8px;">Final Term</th>
                    <th style="border: 1px solid #333; padding: 8px;">Total</th>
                    <th style="border: 1px solid #333; padding: 8px;">Grade</th>
                </tr>
    """
    
    total_marks = 0
    total_subjects = len(student_data['subjects'])
    
    for i, subject in enumerate(student_data['subjects']):
        mid = student_data['mid_marks'][i]
        final = student_data['final_marks'][i]
        
        if subject.lower() == 'computer':
            grade = 'A' if (mid + final) >= 60 else 'B'
            total_str = grade
        else:
            total = mid + final
            total_marks += total
            total_str = str(total)
            
            if total >= 80: grade = 'A+'
            elif total >= 70: grade = 'A'
            elif total >= 60: grade = 'B'
            elif total >= 50: grade = 'C'
            elif total >= 40: grade = 'D'
            else: grade = 'F'
        
        html += f"""
                <tr>
                    <td style="border: 1px solid #333; padding: 8px;">{subject}</td>
                    <td style="border: 1px solid #333; padding: 8px; text-align: center;">{mid}</td>
                    <td style="border: 1px solid #333; padding: 8px; text-align: center;">{final}</td>
                    <td style="border: 1px solid #333; padding: 8px; text-align: center;">{total_str}</td>
                    <td style="border: 1px solid #333; padding: 8px; text-align: center;">{grade}</td>
                </tr>
        """
    
    # Calculate percentage and result
    max_marks = (total_subjects - 1) * 100  # Excluding computer
    percentage = (total_marks / max_marks * 100) if max_marks > 0 else 0
    result = "PASS" if percentage >= 40 else "FAIL"
    
    html += f"""
            </table>
            
            <div style="text-align: left; margin: 20px 0;">
                <p><strong>Total Marks:</strong> {total_marks}/{max_marks}</p>
                <p><strong>Percentage:</strong> {percentage:.1f}%</p>
                <p><strong>Result:</strong> <span style="color: {'green' if result == 'PASS' else 'red'};">{result}</span></p>
            </div>
            
            <div style="margin-top: 40px;">
                <p><strong>Teacher's Remarks:</strong> Keep up the good work!</p>
            </div>
            
            <div style="margin-top: 40px; display: flex; justify-content: space-between;">
                <div>Class Teacher: _______________</div>
                <div>Principal: _______________</div>
            </div>
        </div>
    </div>
    """
    
    return html

def main():
    st.markdown('<h1 class="main-header">🏫 Report Card System</h1>', unsafe_allow_html=True)
    
    # Load students
    students = load_students()
    
    if not students:
        st.error("No student data found. Please check if Excel files are uploaded.")
        return
    
    st.success(f"✅ Loaded {len(students)} students from {len(set(s['class'] for s in students.values()))} classes")
    
    # Sidebar for student selection
    st.sidebar.header("Select Student")
    
    # Class filter
    classes = sorted(set(student['class'] for student in students.values()))
    selected_class = st.sidebar.selectbox("Filter by Class:", ["All Classes"] + classes)
    
    # Filter students by class
    if selected_class == "All Classes":
        filtered_students = students
    else:
        filtered_students = {k: v for k, v in students.items() if v['class'] == selected_class}
    
    # Student selection
    student_options = [f"{student_id} - {data['name']} ({data['class']})" 
                      for student_id, data in filtered_students.items()]
    
    if student_options:
        selected_student = st.sidebar.selectbox("Choose Student:", student_options)
        student_id = selected_student.split(" - ")[0]
        
        if st.sidebar.button("Generate Report Card", type="primary"):
            student_data = students[student_id]
            
            # Display student info
            col1, col2 = st.columns(2)
            with col1:
                st.info(f"**Student:** {student_data['name']}")
            with col2:
                st.info(f"**Class:** {student_data['class']}")
            
            # Generate and display report card
            report_html = generate_report_card(student_data)
            st.markdown(report_html, unsafe_allow_html=True)
            
            # Download button
            st.download_button(
                label="📄 Download Report Card",
                data=report_html,
                file_name=f"Report_Card_{student_id}_{student_data['name']}.html",
                mime="text/html"
            )
    else:
        st.warning("No students found for the selected class.")

if __name__ == "__main__":
    main()
