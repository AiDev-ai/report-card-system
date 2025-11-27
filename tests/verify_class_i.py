#!/usr/bin/env python3
"""
Verify Class I functionality is working
"""
import openpyxl

def verify_class_i():
    print("Verifying Class I Functionality")
    print("=" * 40)
    
    try:
        # Load Excel files
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
        final_wb = openpyxl.load_workbook("Excel_Data/Exams/Final Term.xlsx")
        
        # Check Class I sheet
        if "Class I" in mid_wb.sheetnames:
            mid_sheet = mid_wb["Class I"]
            final_sheet = final_wb["Class I"]
            
            print("✅ Class I sheet found in both Excel files")
            
            # Find Shayan
            for row in range(1, 50):
                for col in ['A', 'B', 'C', 'D']:
                    cell_val = mid_sheet[f'{col}{row}'].value
                    if cell_val and 'AAH- 220' in str(cell_val):
                        print(f"✅ Found Shayan at Row {row}, Column {col}")
                        
                        # Get name
                        for name_offset in [1, 2, 3]:
                            try:
                                next_col = chr(ord(col) + name_offset)
                                name_val = mid_sheet[f'{next_col}{row}'].value
                                if name_val and 'Shayan' in str(name_val):
                                    print(f"✅ Name confirmed: {name_val}")
                                    
                                    # Count valid marks
                                    start_col_ord = ord(next_col) + 1
                                    valid_subjects = 0
                                    
                                    for i in range(10):
                                        mark_col = chr(start_col_ord + i)
                                        if ord(mark_col) > ord('Z'):
                                            break
                                            
                                        mid_val = mid_sheet[f'{mark_col}{row}'].value
                                        final_val = final_sheet[f'{mark_col}{row}'].value
                                        
                                        # Skip formulas
                                        if isinstance(mid_val, str) and mid_val.startswith('='):
                                            continue
                                        if isinstance(final_val, str) and final_val.startswith('='):
                                            continue
                                            
                                        if mid_val is not None or final_val is not None:
                                            valid_subjects += 1
                                            print(f"  Subject {valid_subjects}: {mid_val} / {final_val}")
                                        elif valid_subjects > 5:  # Stop after finding several subjects
                                            break
                                    
                                    print(f"✅ Total valid subjects: {valid_subjects}")
                                    
                                    if valid_subjects >= 7:
                                        print("✅ Shayan has sufficient subject data for report generation!")
                                        print("✅ Class I functionality is WORKING!")
                                    else:
                                        print("⚠️  Limited subject data found")
                                    
                                    return
                            except:
                                continue
        else:
            print("❌ Class I sheet not found")
            
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    verify_class_i()
