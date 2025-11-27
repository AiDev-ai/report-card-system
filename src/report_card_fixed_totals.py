import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
import openpyxl
import subprocess
import os

class ReportCardFixedTotals:
    def __init__(self, root):
        self.root = root
        self.root.title("Alkhidmat Aghosh Hala Auto Report Card Generator")
        self.root.geometry("1200x900")
        self.root.configure(bg='white')
        
        # Force window to appear in taskbar
        self.root.wm_attributes('-topmost', False)
        self.root.lift()
        self.root.focus_force()
        
        self.load_data()
        self.create_interface()
        
        # Set icon after everything is loaded
        self.root.after_idle(self.set_taskbar_icon)
        
        # Auto-refresh timer to check for new data every 30 seconds
        self.setup_auto_refresh()
    
    def set_taskbar_icon(self):
        """Set taskbar icon after window is fully loaded"""
        try:
            # Force taskbar icon update
            self.root.iconbitmap("../assets/icon.ico")
            self.root.update_idletasks()
        except:
            try:
                icon_img = Image.open("../assets/Aghos logo.png")
                icon_img = icon_img.resize((32, 32))
                self.icon_photo = ImageTk.PhotoImage(icon_img)
                self.root.iconphoto(True, self.icon_photo)
                self.root.update_idletasks()
            except:
                pass
    
    def setup_auto_refresh(self):
        """Setup automatic data refresh to detect new students"""
        self.check_for_new_data()
        # Schedule next check in 30 seconds
        self.root.after(30000, self.setup_auto_refresh)
    
    def check_for_new_data(self):
        """Check if new data has been added to Excel files"""
        try:
            # Store current student count
            current_count = len(self.students)
            
            # Reload data silently
            old_students = self.students.copy()
            self.load_data_silent()
            
            # Check if new students were added
            new_count = len(self.students)
            if new_count > current_count:
                new_students = []
                for student_id in self.students:
                    if student_id not in old_students:
                        student_name = self.students[student_id]['name']
                        student_class = self.students[student_id]['class']
                        new_students.append(f"Class {student_class} - {student_id} - {student_name}")
                
                # Update the interface
                self.refresh_student_list()
                
                # Show notification
                messagebox.showinfo("New Data Detected!", 
                    f"✅ Found {new_count - current_count} new student(s):\n\n" + 
                    "\n".join(new_students[:5]) + 
                    (f"\n... and {len(new_students)-5} more" if len(new_students) > 5 else ""))
                    
        except Exception as e:
            # Silent fail - don't interrupt user
            pass
    
    def load_data_silent(self):
        """Load data without showing messages - for background refresh"""
        try:
            mid_wb = openpyxl.load_workbook("../data/Exams/Mid Term.xlsx")
            final_wb = openpyxl.load_workbook("../data/Exams/Final Term.xlsx")
            
            self.students = {}
            self.available_ids = []
            
            # Get all common sheets between both workbooks
            mid_sheets = set(mid_wb.sheetnames)
            final_sheets = set(final_wb.sheetnames)
            common_sheets = mid_sheets.intersection(final_sheets)
            
            # Filter out non-class sheets
            class_sheets = []
            for sheet in common_sheets:
                sheet_lower = sheet.lower()
                if any(keyword in sheet_lower for keyword in ['class', 'prep', 'nursery', 'kg', 'grade']):
                    class_sheets.append(sheet)
            
            for sheet_name in class_sheets:
                mid_sheet = mid_wb[sheet_name]
                final_sheet = final_wb[sheet_name]
                
                # Extract class name more flexibly
                student_class = sheet_name.replace('Class', '').replace('class', '').strip()
                if student_class.startswith('  '):
                    student_class = student_class.strip()
                
                for row in range(1, 200):
                    try:
                        student_id = None
                        student_name = None
                        
                        for col in ['A', 'B', 'C']:
                            cell_val = mid_sheet[f'{col}{row}'].value
                            if cell_val and str(cell_val).startswith('AAH'):
                                next_col = chr(ord(col) + 1)
                                name_val = mid_sheet[f'{next_col}{row}'].value
                                if name_val and not str(name_val).startswith('AAH'):
                                    student_id = str(cell_val).strip()
                                    student_name = str(name_val).strip()
                                    break
                        
                        if student_id and student_name and student_id not in self.students:
                            mid_marks = []
                            final_marks = []
                            
                            # Read marks based on class structure
                            if student_class in ['II', 'III']:
                                mark_cols = ['D', 'E', 'F', 'G', 'H', 'I', 'J']
                            elif student_class in ['IV', 'V', 'VI', 'VII', 'VIII']:
                                mark_cols = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
                            elif student_class in ['IX', 'X']:
                                mark_cols = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
                            else:
                                # For new classes like Prep, Class I - auto-detect columns
                                mark_cols = ['D', 'E', 'F', 'G', 'H', 'I', 'J']
                            
                            for col in mark_cols:
                                mid_val = mid_sheet[f'{col}{row}'].value
                                final_val = final_sheet[f'{col}{row}'].value
                                
                                # Skip formula columns
                                if isinstance(mid_val, str) and mid_val.startswith('='):
                                    mid_val = 0
                                if isinstance(final_val, str) and final_val.startswith('='):
                                    final_val = 0
                                
                                mid_marks.append(mid_val if mid_val is not None else 0)
                                final_marks.append(final_val if final_val is not None else 0)
                            
                            self.students[student_id] = {
                                'name': student_name,
                                'class': student_class,
                                'mid_marks': mid_marks,
                                'final_marks': final_marks
                            }
                            self.available_ids.append(student_id)
                            
                    except:
                        continue
            
            self.sort_students_by_class()
            
        except:
            pass  # Silent fail for background refresh
    
    def refresh_student_list(self):
        """Refresh the student list display"""
        try:
            # Clear current list
            self.student_listbox.delete(0, tk.END)
            
            # Repopulate with updated data
            for student_id in self.available_ids:
                student_name = self.students[student_id]['name']
                student_class = self.students[student_id]['class']
                display_text = f"Class {student_class} - {student_id} - {student_name}"
                self.student_listbox.insert(tk.END, display_text)
                
        except:
            pass
    
    def get_class_total_marks(self, student_class, subjects_data):
        """Dynamically calculate total marks based on actual subjects"""
        
        # For Prep and Class I, calculate based on actual subjects
        if student_class in ['Prep', 'I']:
            # Get the actual subjects for this class
            subjects = self.get_subjects(student_class)
            subject_count = len(subjects)
            
            # Prep typically has 5 subjects × 100 = 500
            # Class I typically has 7 subjects but may vary
            if student_class == 'Prep':
                return subject_count * 100  # 5 subjects = 500 marks
            else:  # Class I
                return subject_count * 100  # 7 subjects = 700 marks
        
        # Default totals for other known classes
        class_defaults = {
            'II': 500, 'III': 500,
            'IV': 550, 'V': 550, 'VI': 550, 'VII': 550, 'VIII': 550,
            'IX': 650, 'X': 650
        }
        
        # If we have a default, use it
        if student_class in class_defaults:
            return class_defaults[student_class]
        
        # Otherwise, calculate from actual subject marks
        total = 0
        for subject_data in subjects_data:
            if len(subject_data) >= 2 and subject_data[1]:  # Has mid term marks
                try:
                    # Estimate subject total (usually mid + final = total, so mid*2 as estimate)
                    mid_mark = float(subject_data[1])
                    if mid_mark <= 50:  # Likely out of 50
                        total += 50
                    else:  # Likely out of 100
                        total += 100
                except:
                    total += 100  # Default assumption
        
        return total if total > 0 else 500  # Fallback to 500
    
    def get_remarks_by_percentage(self, percentage):
        """Get remarks based on obtained percentage"""
        if percentage >= 85:
            return "Outstanding"
        elif percentage >= 80:
            return "Excellent"
        elif percentage >= 70:
            return "Very Good"
        elif percentage >= 60:
            return "Good"
        elif percentage >= 50:
            return "Satisfactory"
        elif percentage >= 40:
            return "Fair"
        else:
            return "Needs Attention"
    
    def sort_students_by_class(self):
        """Sort students by class order, then by ID"""
        # Define class order
        class_order = {
            'Prep': 0, 'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5,
            'VI': 6, 'VII': 7, 'VIII': 8, 'IX': 9, 'X': 10
        }
        
        def get_sort_key(student_id):
            student_class = self.students[student_id]['class']
            class_priority = class_order.get(student_class, 99)
            return (class_priority, student_id)
        
        self.available_ids.sort(key=get_sort_key)
    
    def get_logo_base64(self):
        try:
            import base64
            logo_path = "../assets/Aghos logo.png"
            with open(logo_path, "rb") as img_file:
                return base64.b64encode(img_file.read()).decode('utf-8')
        except:
            return None
    
    def load_data(self):
        try:
            mid_wb = openpyxl.load_workbook("../data/Exams/Mid Term.xlsx")
            final_wb = openpyxl.load_workbook("../data/Exams/Final Term.xlsx")
            
            self.students = {}
            self.available_ids = []
            
            # Get all common sheets between both workbooks
            mid_sheets = set(mid_wb.sheetnames)
            final_sheets = set(final_wb.sheetnames)
            common_sheets = mid_sheets.intersection(final_sheets)
            
            # Filter out non-class sheets (like summary, etc.)
            class_sheets = []
            for sheet in common_sheets:
                sheet_lower = sheet.lower()
                if any(keyword in sheet_lower for keyword in ['class', 'prep', 'nursery', 'kg', 'grade']):
                    class_sheets.append(sheet)
            
            print(f"Found {len(class_sheets)} class sheets: {class_sheets}")
            
            for sheet_name in class_sheets:
                mid_sheet = mid_wb[sheet_name]
                final_sheet = final_wb[sheet_name]
                
                # Extract class name more flexibly
                student_class = sheet_name.replace('Class', '').replace('class', '').strip()
                if student_class.startswith('  '):
                    student_class = student_class.strip()
                
                for row in range(1, 300):  # Increased range
                    try:
                        student_id = None
                        student_name = None
                        id_col = None
                        name_col = None
                        
                        # Enhanced student ID detection - check more columns
                        for col in ['A', 'B', 'C', 'D']:
                            cell_val = mid_sheet[f'{col}{row}'].value
                            if cell_val and str(cell_val).strip().startswith('AAH'):
                                # Look for name in next few columns
                                for name_offset in [1, 2, 3]:
                                    try:
                                        next_col = chr(ord(col) + name_offset)
                                        name_val = mid_sheet[f'{next_col}{row}'].value
                                        if name_val and not str(name_val).strip().startswith('AAH') and len(str(name_val).strip()) > 2:
                                            student_id = str(cell_val).strip()
                                            student_name = str(name_val).strip()
                                            id_col = col
                                            name_col = next_col
                                            break
                                    except:
                                        continue
                                if student_id and student_name:
                                    break
                        
                        if student_id and student_name and student_id not in self.students:
                            mid_marks = []
                            final_marks = []
                            subject_names = []
                            
                            # Determine mark columns based on where name ends
                            if name_col:
                                start_col_ord = ord(name_col) + 1
                            else:
                                start_col_ord = ord('D')  # Default fallback
                            
                            # Only find subject headers for Prep class (since it's different)
                            subject_names = []
                            if student_class == 'Prep':
                                for header_row in range(1, 10):  # Check first 10 rows
                                    potential_subjects = []
                                    for i in range(15):
                                        col = chr(start_col_ord + i)
                                        if ord(col) > ord('Z'):
                                            break
                                        header_cell = mid_sheet[f'{col}{header_row}'].value
                                        if header_cell and isinstance(header_cell, str):
                                            header_name = str(header_cell).strip()
                                            # Check if this looks like a subject name
                                            if (len(header_name) > 2 and 
                                                not header_name.startswith('AAH') and 
                                                not header_name.startswith('=') and
                                                not header_name.isdigit() and
                                                header_name not in ['Total', 'Secured Ma', 'Consolidat', 'Aggregate', 'Secured Marks', 'Remarks']):
                                                potential_subjects.append(header_name)
                                    
                                    # If we found multiple subjects in this row, use them
                                    if len(potential_subjects) >= 3:
                                        subject_names = potential_subjects
                                        print(f"Found Prep subjects in row {header_row}: {subject_names}")
                                        break
                            
                            # Read marks from columns after the name
                            marks_found = False
                            for i in range(15):  # Check up to 15 subjects
                                col = chr(start_col_ord + i)
                                if ord(col) > ord('Z'):  # Don't go beyond column Z
                                    break
                                    
                                try:
                                    mid_val = mid_sheet[f'{col}{row}'].value
                                    final_val = final_sheet[f'{col}{row}'].value
                                    
                                    # Skip formula columns (like totals, percentages)
                                    if isinstance(mid_val, str) and mid_val.startswith('='):
                                        continue
                                    if isinstance(final_val, str) and final_val.startswith('='):
                                        continue
                                    
                                    # Check if this column has actual mark data
                                    if mid_val is not None or final_val is not None:
                                        # Convert to appropriate type, handle grades
                                        if isinstance(mid_val, (int, float)):
                                            mid_marks.append(float(mid_val))
                                        elif isinstance(mid_val, str) and mid_val.strip().replace('.','').isdigit():
                                            mid_marks.append(float(mid_val))
                                        elif isinstance(mid_val, str) and mid_val.strip() in ['A+', 'A', 'B', 'C', 'D', 'E', 'F']:
                                            mid_marks.append(mid_val.strip())  # Keep grades as text
                                        else:
                                            mid_marks.append(mid_val if mid_val is not None else 0)
                                            
                                        if isinstance(final_val, (int, float)):
                                            final_marks.append(float(final_val))
                                        elif isinstance(final_val, str) and final_val.strip().replace('.','').isdigit():
                                            final_marks.append(float(final_val))
                                        elif isinstance(final_val, str) and final_val.strip() in ['A+', 'A', 'B', 'C', 'D', 'E', 'F']:
                                            final_marks.append(final_val.strip())  # Keep grades as text
                                        else:
                                            final_marks.append(final_val if final_val is not None else 0)
                                            
                                        marks_found = True
                                    elif marks_found and i > 7:  # Stop after finding some marks and hitting empty columns
                                        break
                                except:
                                    continue
                            
                            # Only add student if we found marks data
                            if marks_found and (mid_marks or final_marks):
                                self.students[student_id] = {
                                    'name': student_name,
                                    'class': student_class,
                                    'mid_marks': mid_marks,
                                    'final_marks': final_marks,
                                    'subject_names': subject_names  # Store actual subject names
                                }
                                self.available_ids.append(student_id)
                                print(f"Added: {student_id} - {student_name} (Class {student_class}) - {len(mid_marks)} subjects")
                                if subject_names:
                                    print(f"  Subjects: {subject_names[:len(mid_marks)]}")
                                
                    except Exception as e:
                        continue
            
            self.sort_students_by_class()  # Sort by class instead of alphabetically
            print(f"Loaded {len(self.students)} students")
            
        except Exception as e:
            messagebox.showerror("Error", f"Data load error: {e}")
    
    def create_interface(self):
        main_frame = tk.Frame(self.root, bg='white')
        main_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Left panel
        left_frame = tk.Frame(main_frame, bg='lightgray', width=280)
        left_frame.pack(side='left', fill='y', padx=5)
        left_frame.pack_propagate(False)
        
        tk.Label(left_frame, text=f"All {len(self.students)} Students", font=('Arial', 12, 'bold'), bg='lightgray').pack(pady=5)
        
        search_frame = tk.Frame(left_frame, bg='lightgray')
        search_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(search_frame, text="Search:", font=('Arial', 10), bg='lightgray').pack(anchor='w')
        self.search_entry = tk.Entry(search_frame, font=('Arial', 10))
        self.search_entry.pack(fill='x', pady=2)
        self.search_entry.bind('<KeyRelease>', self.filter_students)
        
        listbox_frame = tk.Frame(left_frame, bg='lightgray')
        listbox_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        scrollbar_list = ttk.Scrollbar(listbox_frame)
        scrollbar_list.pack(side='right', fill='y')
        
        self.student_listbox = tk.Listbox(listbox_frame, font=('Arial', 9), yscrollcommand=scrollbar_list.set)
        self.student_listbox.pack(side='left', fill='both', expand=True)
        scrollbar_list.config(command=self.student_listbox.yview)
        
        for student_id in self.available_ids:
            student_name = self.students[student_id]['name']
            student_class = self.students[student_id]['class']
            display_text = f"Class {student_class} - {student_id} - {student_name}"
            self.student_listbox.insert(tk.END, display_text)
        
        self.student_listbox.bind('<Double-Button-1>', self.on_student_select)
        
        # Right panel
        right_frame = tk.Frame(main_frame, bg='white')
        right_frame.pack(side='right', fill='both', expand=True, padx=5)
        
        control_frame = tk.Frame(right_frame, bg='lightblue', height=50)
        control_frame.pack(fill='x', pady=5)
        control_frame.pack_propagate(False)
        
        tk.Button(control_frame, text="Show Selected", command=self.show_selected_report, font=('Arial', 10, 'bold')).pack(side='left', padx=10, pady=10)
        tk.Button(control_frame, text="🔄 Refresh Data", command=self.refresh_data, font=('Arial', 10, 'bold'), bg='orange', fg='white').pack(side='left', padx=5, pady=10)
        tk.Button(control_frame, text="🖨️ Direct Print", command=self.direct_print, font=('Arial', 10, 'bold'), bg='blue', fg='white').pack(side='left', padx=5, pady=10)
        tk.Button(control_frame, text="💾 Save HTML", command=self.print_report, font=('Arial', 10, 'bold'), bg='green', fg='white').pack(side='left', padx=5, pady=10)
        
        self.current_student_label = tk.Label(control_frame, text="Select a student from the list", font=('Arial', 10), bg='lightblue')
        self.current_student_label.pack(side='right', padx=10, pady=10)
        
        # Report card area
        self.canvas = tk.Canvas(right_frame, bg='white')
        scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=self.canvas.yview)
        self.report_frame = tk.Frame(self.canvas, bg='white', relief='solid', borderwidth=2)
        
        self.report_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.report_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.create_report()
    
    def filter_students(self, event=None):
        search_term = self.search_entry.get().lower()
        self.student_listbox.delete(0, tk.END)
        
        for student_id in self.available_ids:
            student_name = self.students[student_id]['name']
            student_class = self.students[student_id]['class']
            display_text = f"Class {student_class} - {student_id} - {student_name}"
            
            if search_term in student_id.lower() or search_term in student_name.lower() or search_term in student_class.lower():
                self.student_listbox.insert(tk.END, display_text)
    
    def on_student_select(self, event=None):
        self.show_selected_report()
    
    def show_selected_report(self):
        selection = self.student_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a student from the list")
            return
        
        selected_text = self.student_listbox.get(selection[0])
        # Extract student ID from "Class X - AAH001 - Student Name" format
        parts = selected_text.split(' - ')
        if len(parts) >= 2:
            student_id = parts[1]  # AAH001 is the second part
        else:
            student_id = parts[0]  # Fallback
        
        student = self.students.get(student_id)
        if not student:
            messagebox.showerror("Error", "Student data not found")
            return
        
        self.current_student_id = student_id
        self.current_student_label.config(text=f"Showing: {student_id} - {student['name']}")
        
        # Clear previous table
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        
        # Update student info
        self.name_label.config(text=f"{student['name'] or ''}", font=('Arial', 11, 'bold'))
        self.class_label.config(text=f"{student['class'] or ''}", font=('Arial', 11, 'bold'))
        self.id_label.config(text=f"{student_id}", font=('Arial', 11, 'bold'))
        
        subjects = self.get_subjects(student['class'], student_id)
        self.create_dynamic_table(subjects, student)
    
    def get_subjects(self, cls, student_id=None):
        """Get subjects - use predefined for most classes, dynamic only for Prep"""
        
        # For Prep class, use dynamic detection since it's different
        if cls == 'Prep':
            # First try to get subjects from stored student data
            if student_id and student_id in self.students:
                stored_subjects = self.students[student_id].get('subject_names', [])
                if stored_subjects:
                    # Remove non-subject columns and limit to actual subjects
                    actual_subjects = []
                    for subj in stored_subjects:
                        if subj not in ['Secured Marks', 'Remarks', 'Total', 'Aggregate']:
                            actual_subjects.append(subj)
                    if actual_subjects:
                        return actual_subjects[:5]  # Prep typically has 5 subjects
            
            # Fallback for Prep
            return ['English', 'Urdu', 'Mathematics', 'Islam/GK', 'Art']
        
        # For all other classes, use predefined subjects (the original working system)
        elif cls in ['I']:
            return ['English', 'Urdu', 'GK', 'Mathematics', 'Computer', 'Islamiat', 'Sindhi']
        elif cls in ['II', 'III']:
            return ['English', 'Urdu', 'GK', 'Mathematics', 'Computer', 'Islamiat', 'Sindhi']
        elif cls in ['IV', 'V', 'VI', 'VII', 'VIII']:
            return ['English', 'Urdu', 'Science', 'Mathematics', 'Computer', 'Social Studies', 'Islamiat', 'Sindhi']
        elif cls == 'IX':
            return ['English', 'Urdu', 'Mathematics', 'Biology', 'Physics', 'Chemistry', 'Computer', 'Islamiat']
        elif cls == 'X':
            return ['English', 'Urdu', 'Mathematics', 'Biology', 'Physics', 'Chemistry', 'Computer', 'Pakistan Studies']
        
        return []
    
    def get_total_marks(self, cls):
        # Dynamic calculation for Prep and Class I
        if cls in ['Prep', 'I']:
            subjects = self.get_subjects(cls)
            return len(subjects) * 100  # Each subject worth 100 marks
        elif cls in ['II', 'III']:
            return 700  # 7 subjects × 100 each (including Computer)
        elif cls in ['IV', 'V', 'VI', 'VII', 'VIII']:
            return 800  # 8 subjects × 100 each (including Computer)
        elif cls == 'IX':
            return 800  # 8 subjects × 100 each (including Computer)
        elif cls == 'X':
            return 800  # 8 subjects × 100 each (including Computer)
        return 0
    
    def is_computer_subject(self, subject):
        return subject == 'Computer'
    
    def get_computer_grade(self, student, subject_index, term='both'):
        try:
            # Computer is at the correct subject_index position
            computer_index = subject_index
            
            mid_val = student['mid_marks'][computer_index] if computer_index < len(student['mid_marks']) else None
            final_val = student['final_marks'][computer_index] if computer_index < len(student['final_marks']) else None
            
            if term == 'mid':
                # Check if it's already a grade (string)
                if isinstance(mid_val, str) and mid_val.strip() in ['A+', 'A', 'B', 'C', 'D', 'E', 'F']:
                    return mid_val.strip()
                # If it's a number, convert to grade
                elif isinstance(mid_val, (int, float)) and mid_val > 0:
                    return self.percentage_to_grade(mid_val)
                else:
                    return 'C'
                    
            elif term == 'final':
                # Check if it's already a grade (string)
                if isinstance(final_val, str) and final_val.strip() in ['A+', 'A', 'B', 'C', 'D', 'E', 'F']:
                    return final_val.strip()
                # If it's a number, convert to grade
                elif isinstance(final_val, (int, float)) and final_val > 0:
                    return self.percentage_to_grade(final_val)
                else:
                    return 'C'
            else:
                # Return average grade for both terms
                mid_grade = self.get_computer_grade(student, subject_index, 'mid')
                final_grade = self.get_computer_grade(student, subject_index, 'final')
                # If both grades are the same, return that grade
                if mid_grade == final_grade:
                    return mid_grade
                # Otherwise return the better grade
                return self.get_better_grade(mid_grade, final_grade)
        except Exception as e:
            print(f"Error getting computer grade for student {student.get('name', 'Unknown')}: {e}")
            return 'C'
    
    def grade_to_percentage(self, grade):
        """Convert grade to percentage for calculation"""
        grade_map = {
            'A+': 95, 'A': 85, 'B': 75, 'C': 65, 
            'D': 55, 'E': 45, 'F': 25
        }
        return grade_map.get(grade, 65)
    
    def percentage_to_grade(self, percentage):
        """Convert percentage to grade"""
        if percentage >= 91: return 'A+'
        elif percentage >= 80: return 'A'
        elif percentage >= 70: return 'B'
        elif percentage >= 60: return 'C'
        elif percentage >= 50: return 'D'
        elif percentage >= 35: return 'E'
        else: return 'F'
    
    def get_better_grade(self, grade1, grade2):
        """Return the better of two grades"""
        grade_order = ['F', 'E', 'D', 'C', 'B', 'A', 'A+']
        try:
            index1 = grade_order.index(grade1)
            index2 = grade_order.index(grade2)
            return grade1 if index1 > index2 else grade2
        except:
            return grade1
    
    def get_remarks_for_grade(self, grade):
        """Get remarks based on grade using percentage conversion"""
        # Convert grade to percentage and use percentage-based remarks
        percentage = self.grade_to_percentage(grade)
        return self.get_remarks_by_percentage(percentage)
    
    def create_dynamic_table(self, subjects, student):
        headers = [
            ('Sr. #', 6), ('Subjects', 16), ('1st Term\n100/50', 10), ('Weighted\n20%', 9),
            ('2nd Term\n100/50', 10), ('Weighted\n80%', 9), ('Aggregate', 10), 
            ('Percentage', 10), ('Grade', 8), ('Remarks', 12)
        ]
        
        for i, (header, width) in enumerate(headers):
            label = tk.Label(self.table_frame, text=header, font=('Arial', 9, 'bold'), 
                           relief='solid', borderwidth=1, bg='lightgray', width=width, 
                           height=3, wraplength=width*8, justify='center')
            label.grid(row=0, column=i, sticky='nsew', padx=1, pady=1)
        
        for i in range(10):
            self.table_frame.grid_columnconfigure(i, weight=1)
        
        total_mid = 0
        total_final = 0
        total_agg = 0
        calculation_subjects = 0
        
        for i, subject in enumerate(subjects):
            row_num = i + 1
            
            if self.is_computer_subject(subject):
                # Computer subject - show grades but calculate differently
                mid_grade = self.get_computer_grade(student, i, 'mid')
                final_grade = self.get_computer_grade(student, i, 'final')
                
                # Convert grades to percentage for calculation
                mid_perc = self.grade_to_percentage(mid_grade)
                final_perc = self.grade_to_percentage(final_grade)
                
                # Calculate average of both grades
                avg_perc = (mid_perc + final_perc) / 2
                avg_grade = self.percentage_to_grade(avg_perc)
                
                remarks = self.get_remarks_for_grade(avg_grade)
                
                row_data = [
                    str(i+1), subject, mid_grade, '-', 
                    final_grade, '-', avg_grade, 
                    f"{avg_perc:.0f}", avg_grade, remarks
                ]
                
                total_mid += mid_perc
                total_final += final_perc
                total_agg += avg_perc
                calculation_subjects += 1
            else:
                # Regular subjects with marks
                mid = student['mid_marks'][i] if i < len(student['mid_marks']) else 0
                final = student['final_marks'][i] if i < len(student['final_marks']) else 0
                
                # Determine max marks for this subject based on class and subject
                if student['class'] in ['II', 'III']:
                    # English, Urdu, GK, Maths = 100 marks; Islamiat, Sindhi = 50 marks
                    max_marks = 100 if subject in ['English', 'Urdu', 'GK', 'Mathematics'] else 50
                elif student['class'] in ['IV', 'V', 'VI', 'VII', 'VIII']:
                    # English, Urdu, Science, Maths = 100 marks; S.S, Islamiat, Sindhi = 50 marks
                    max_marks = 100 if subject in ['English', 'Urdu', 'Science', 'Mathematics'] else 50
                elif student['class'] in ['IX', 'X']:
                    # English, Urdu, Maths, Bio, Phy, Che = 100 marks; Islamiat/PS = 50 marks
                    max_marks = 100 if subject in ['English', 'Urdu', 'Mathematics', 'Biology', 'Physics', 'Chemistry'] else 50
                else:
                    max_marks = 100
                
                w_mid = (mid * 20) / 100
                w_final = (final * 80) / 100
                agg = w_mid + w_final
                perc = (agg / max_marks) * 100  # Calculate percentage based on max marks
                
                grade = 'A+' if perc >= 91 else 'A' if perc >= 80 else 'B' if perc >= 70 else 'C' if perc >= 60 else 'D' if perc >= 50 else 'E' if perc >= 35 else 'F'
                remarks = self.get_remarks_for_grade(grade)
                
                row_data = [
                    str(i+1), subject, str(mid), f"{w_mid:.1f}", 
                    str(final), f"{w_final:.1f}", f"{agg:.0f}", 
                    f"{perc:.0f}", grade, remarks
                ]
                
                total_mid += mid
                total_final += final
                total_agg += agg
                calculation_subjects += 1
            
            for j, (data, (_, width)) in enumerate(zip(row_data, headers)):
                label = tk.Label(self.table_frame, text=data, font=('Arial', 9), 
                               relief='solid', borderwidth=1, bg='white', width=width, 
                               height=2, wraplength=width*7, justify='center', anchor='center')
                label.grid(row=row_num, column=j, sticky='nsew', padx=1, pady=1)
        
        # Total row
        total_row_num = len(subjects) + 1
        
        # Calculate correct total marks dynamically
        class_total_marks = self.get_class_total_marks(student['class'], subjects)
        
        total_data = [
            '', f'Total [{class_total_marks}]', f'{total_mid}', '', 
            f'{total_final}', '', f'{total_agg:.0f}', 
            f'{total_agg/calculation_subjects:.0f}' if calculation_subjects > 0 else '0', '', ''
        ]
        
        for j, (data, (_, width)) in enumerate(zip(total_data, headers)):
            bg = 'lightgray' if j in [0,1,2,4,6,7] else 'white'
            font_weight = 'bold' if j in [1,2,4,6,7] else 'normal'
            label = tk.Label(self.table_frame, text=data, font=('Arial', 9, font_weight), 
                           relief='solid', borderwidth=1, bg=bg, width=width, 
                           height=2, wraplength=width*7, justify='center', anchor='center')
            label.grid(row=total_row_num, column=j, sticky='nsew', padx=1, pady=1)
        
        # Update footer
        overall_perc = total_agg/calculation_subjects if calculation_subjects > 0 else 0
        overall_grade = 'A+' if overall_perc >= 91 else 'A' if overall_perc >= 80 else 'B' if overall_perc >= 70 else 'C'
        pass_fail = "Pass" if overall_perc >= 40 else "Fail"
        
        self.overall_label.config(text=f"{overall_perc:.0f}%")
        self.grade_label.config(text=overall_grade)
        self.pass_fail_label.config(text=pass_fail)
    
    def refresh_data(self):
        """Refresh student data from Excel files"""
        try:
            # Show loading message
            self.current_student_label.config(text="🔄 Refreshing data...")
            self.root.update()
            
            # Reload data
            old_count = len(self.students)
            self.load_data()
            new_count = len(self.students)
            
            # Update student list
            self.student_listbox.delete(0, tk.END)
            for student_id in self.available_ids:
                student_name = self.students[student_id]['name']
                student_class = self.students[student_id]['class']
                display_text = f"Class {student_class} - {student_id} - {student_name}"
                self.student_listbox.insert(tk.END, display_text)
            
            # Update labels
            left_frame_label = None
            for widget in self.root.winfo_children():
                if isinstance(widget, tk.Frame):
                    for child in widget.winfo_children():
                        if isinstance(child, tk.Frame):
                            for grandchild in child.winfo_children():
                                if isinstance(grandchild, tk.Label) and "Students" in str(grandchild.cget('text')):
                                    grandchild.config(text=f"All {len(self.students)} Students")
                                    break
            
            # Show result message
            if new_count > old_count:
                added = new_count - old_count
                messagebox.showinfo("Data Refreshed", f"✅ Data refreshed successfully!\n\n📊 Total students: {new_count}\n🆕 New students added: {added}")
            elif new_count < old_count:
                removed = old_count - new_count
                messagebox.showinfo("Data Refreshed", f"✅ Data refreshed successfully!\n\n📊 Total students: {new_count}\n🗑️ Students removed: {removed}")
            else:
                messagebox.showinfo("Data Refreshed", f"✅ Data refreshed successfully!\n\n📊 Total students: {new_count}\n📝 No changes detected")
            
            self.current_student_label.config(text="Select a student from the list")
            
        except Exception as e:
            messagebox.showerror("Refresh Error", f"❌ Error refreshing data:\n{str(e)}")
            self.current_student_label.config(text="Refresh failed - Select a student")
    
    def direct_print(self):
        """Simple working print function"""
        if not hasattr(self, 'current_student_id'):
            messagebox.showwarning("Warning", "Please select a student first")
            return
        
        try:
            student = self.students[self.current_student_id]
            subjects = self.get_subjects(student['class'], self.current_student_id)
            
            # Create HTML content
            html_content = self.create_print_html(student, subjects)
            
            # Save to file
            import tempfile
            temp_file = tempfile.mktemp(suffix='.html')
            with open(temp_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # Open in browser
            import webbrowser
            webbrowser.open(f'file://{temp_file}')
            
            messagebox.showinfo("Print Ready", "Report card opened in browser. Use Ctrl+P to print.")
            
        except Exception as e:
            messagebox.showerror("Print Error", f"Could not print: {str(e)}")
    
    def get_subject_remarks(self, mid_mark, final_mark):
        """Get remarks for individual subject based on marks"""
        try:
            # Calculate average for the subject
            if isinstance(mid_mark, str) and mid_mark in ['A+', 'A', 'B', 'C', 'D', 'E', 'F']:
                # Handle grade-based subjects (like Computer)
                grade_to_perc = {'A+': 95, 'A': 85, 'B': 75, 'C': 65, 'D': 55, 'E': 45, 'F': 35}
                mid_perc = grade_to_perc.get(mid_mark, 50)
            else:
                mid_perc = float(mid_mark) if mid_mark else 0
                
            if isinstance(final_mark, str) and final_mark in ['A+', 'A', 'B', 'C', 'D', 'E', 'F']:
                grade_to_perc = {'A+': 95, 'A': 85, 'B': 75, 'C': 65, 'D': 55, 'E': 45, 'F': 35}
                final_perc = grade_to_perc.get(final_mark, 50)
            else:
                final_perc = float(final_mark) if final_mark else 0
            
            avg_perc = (mid_perc + final_perc) / 2
            
            # Apply the same criteria as overall remarks
            if avg_perc >= 85:
                return "Outstanding"
            elif avg_perc >= 80:
                return "Excellent"
            elif avg_perc >= 70:
                return "Very Good"
            elif avg_perc >= 60:
                return "Good"
            elif avg_perc >= 50:
                return "Satisfactory"
            elif avg_perc >= 40:
                return "Fair"
            else:
                return "Needs Attention"
                
        except:
            return "Fair"
        if not hasattr(self, 'current_student_id'):
            messagebox.showwarning("Warning", "Please select a student first")
            return
        
        try:
            student = self.students[self.current_student_id]
            subjects = self.get_subjects(student['class'], self.current_student_id)
            
            # Create HTML for direct printing
            html_content = self.create_print_html(student, subjects)
            
            # Create HTML file in Windows-accessible location
            import tempfile
            import os
            import re
            
            # Clean student ID for filename
            clean_id = re.sub(r'[^\w\-_]', '_', self.current_student_id)
            
            # Create in Windows temp directory that's accessible
            windows_temp_dir = "/mnt/c/Users/Admin/Documents/temp"
            os.makedirs(windows_temp_dir, exist_ok=True)
            
            temp_file = os.path.join(windows_temp_dir, f"report_card_{clean_id}.html")
            
            # Also save to Output_Files folder for easy access
            output_dir = "../Output_Files"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            permanent_file = os.path.join(output_dir, f"Report_Card_{clean_id}.html")
            
            # Write to both locations
            for file_path in [temp_file, permanent_file]:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(html_content)
            
            # Open in browser using simple Windows command
            try:
                # Convert to Windows path
                windows_path = temp_file.replace('/mnt/c/', 'C:\\').replace('/', '\\')
                # Use rundll32 to open with default browser
                subprocess.run(['rundll32.exe', 'url.dll,FileProtocolHandler', windows_path], check=True)
            except:
                # Fallback: show file location
                messagebox.showinfo("Print Ready", 
                    f"Report card saved to:\n{permanent_file}\n\nDouble-click this file to open in browser and print.")
            
        except Exception as e:
            messagebox.showerror("Print Error", f"Could not open print view: {str(e)}")
    
    def create_print_html(self, student, subjects):
        # Calculate totals
        total_mid = total_final = total_agg = calculation_subjects = 0
        rows_html = ""
        
        for i, subject in enumerate(subjects):
            if self.is_computer_subject(subject):
                mid_grade = self.get_computer_grade(student, i, 'mid')
                final_grade = self.get_computer_grade(student, i, 'final')
                avg_grade = mid_grade if mid_grade == final_grade else self.get_better_grade(mid_grade, final_grade)
                remarks = self.get_remarks_for_grade(avg_grade)
                
                rows_html += f"""
                <tr>
                    <td>{i+1}</td>
                    <td>{subject}</td>
                    <td>{mid_grade}</td>
                    <td>-</td>
                    <td>{final_grade}</td>
                    <td>-</td>
                    <td>{avg_grade}</td>
                    <td>-</td>
                    <td>{avg_grade}</td>
                    <td>{remarks}</td>
                </tr>
                """
            else:
                mid = student['mid_marks'][i] if i < len(student['mid_marks']) else 0
                final = student['final_marks'][i] if i < len(student['final_marks']) else 0
                
                if student['class'] in ['II', 'III']:
                    max_marks = 100 if subject in ['English', 'Urdu', 'GK', 'Mathematics'] else 50
                elif student['class'] in ['IV', 'V', 'VI', 'VII', 'VIII']:
                    max_marks = 100 if subject in ['English', 'Urdu', 'Science', 'Mathematics'] else 50
                elif student['class'] in ['IX', 'X']:
                    max_marks = 100 if subject in ['English', 'Urdu', 'Mathematics', 'Biology', 'Physics', 'Chemistry'] else 50
                else:
                    max_marks = 100
                
                w_mid = (mid * 20) / 100
                w_final = (final * 80) / 100
                agg = w_mid + w_final
                perc = (agg / max_marks) * 100
                
                grade = 'A+' if perc >= 91 else 'A' if perc >= 80 else 'B' if perc >= 70 else 'C' if perc >= 60 else 'D' if perc >= 50 else 'E' if perc >= 35 else 'F'
                remarks = self.get_remarks_for_grade(grade)
                
                rows_html += f"""
                <tr>
                    <td>{i+1}</td>
                    <td>{subject}</td>
                    <td>{mid}</td>
                    <td>{w_mid:.1f}</td>
                    <td>{final}</td>
                    <td>{w_final:.1f}</td>
                    <td>{agg:.0f}</td>
                    <td>{perc:.0f}</td>
                    <td>{grade}</td>
                    <td>{remarks}</td>
                </tr>
                """
                
                total_mid += mid
                total_final += final
                total_agg += agg
                calculation_subjects += 1
        
        # Calculate totals dynamically
        class_total_marks = self.get_class_total_marks(student['class'], subjects)
        
        overall_perc = total_agg/calculation_subjects if calculation_subjects > 0 else 0
        overall_grade = 'A+' if overall_perc >= 91 else 'A' if overall_perc >= 80 else 'B' if overall_perc >= 70 else 'C'
        pass_fail = "Pass" if overall_perc >= 40 else "Fail"
        
        teacher_remarks = self.teacher_remarks.get(1.0, tk.END).strip() if hasattr(self, 'teacher_remarks') else "Student has shown good performance."
        
        # Add automatic performance remarks based on percentage
        performance_remark = self.get_remarks_by_percentage(overall_perc)
        combined_remarks = f"Performance: {performance_remark}. {teacher_remarks}"
        
        # Get base64 logo for embedding
        logo_base64 = self.get_logo_base64()
        logo_src = f"data:image/png;base64,{logo_base64}" if logo_base64 else ""
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Report Card - {self.current_student_id}</title>
            <style>
                @page {{ 
                    size: A4 portrait; 
                    margin: 15mm; 
                }}
                body {{ 
                    font-family: Arial, sans-serif; 
                    margin: 0; 
                    padding: 20px;
                    border: 3px solid black;
                    min-height: 95vh;
                }}
                .header {{ 
                    border-bottom: 2px solid black; 
                    padding: 15px; 
                    margin-bottom: 15px;
                    display: flex;
                    align-items: center;
                }}
                .logo {{ 
                    width: 120px; 
                    height: 120px; 
                    border: 2px solid black;
                    margin-right: 20px;
                    flex-shrink: 0;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    font-size: 14px;
                    font-weight: bold;
                }}
                .header-text {{
                    flex: 1;
                    text-align: center;
                }}
                .header-text h2 {{ 
                    margin: 8px 0; 
                    font-size: 18px; 
                    white-space: nowrap;
                    overflow: visible;
                }}
                .header-text p {{ margin: 5px 0; font-size: 11px; }}
                .header-text h3 {{ margin: 10px 0; font-size: 14px; }}
                
                .student-info {{ 
                    border: 1px solid black; 
                    padding: 10px; 
                    margin-bottom: 15px; 
                    display: flex;
                    justify-content: space-between;
                    font-weight: bold;
                }}
                
                table {{ 
                    width: 100%; 
                    border-collapse: collapse; 
                    margin-bottom: 15px; 
                    font-size: 10px;
                }}
                th, td {{ 
                    border: 1px solid black; 
                    padding: 4px; 
                    text-align: center; 
                }}
                th {{ 
                    background-color: lightgray; 
                    font-weight: bold; 
                    font-size: 9px;
                }}
                .total-row {{ 
                    background-color: lightgray; 
                    font-weight: bold; 
                }}
                
                .footer {{ 
                    display: flex; 
                    margin: 15px 0; 
                    gap: 10px;
                }}
                .grading {{ 
                    border: 1px solid black; 
                    padding: 8px; 
                    flex: 1; 
                    font-size: 9px;
                }}
                .results {{ 
                    border: 1px solid black; 
                    padding: 8px; 
                    flex: 1; 
                    font-size: 10px;
                }}
                
                .teacher-remarks {{ 
                    border: 1px solid black; 
                    padding: 8px; 
                    margin: 10px 0; 
                    min-height: 40px;
                    font-size: 10px;
                }}
                
                .signatures {{ 
                    margin-top: 50px; 
                    display: flex; 
                    justify-content: space-between; 
                }}
                .signature {{ 
                    text-align: center; 
                    font-size: 10px;
                }}
                .signature-line {{ 
                    border-bottom: 1px solid black; 
                    width: 120px; 
                    margin-bottom: 5px; 
                }}
                
                .print-button {{
                    text-align: center; 
                    margin: 20px 0; 
                    padding: 15px; 
                    background: #f0f0f0; 
                    border: 2px solid #333;
                    border-radius: 5px;
                }}
                .print-btn {{
                    background: #4CAF50;
                    color: white;
                    padding: 12px 24px;
                    font-size: 16px;
                    border: none;
                    border-radius: 4px;
                    cursor: pointer;
                    margin: 0 10px;
                }}
                .print-btn:hover {{
                    background: #45a049;
                }}
                
                @media print {{ 
                    body {{ margin: 0; padding: 15px; }}
                    .no-print {{ display: none !important; }}
                    .print-button {{ display: none !important; }}
                }}
            </style>
            <script>
                function printReport() {{
                    window.print();
                }}
                function closeWindow() {{
                    window.close();
                }}
            </script>
        </head>
        <body>
            <div class="print-button no-print">
                <h3>Report Card Ready for Printing</h3>
                <button class="print-btn" onclick="printReport()">🖨️ Print Report Card</button>
                <button class="print-btn" onclick="closeWindow()" style="background: #f44336;">❌ Close</button>
                <p><em>You can also use Ctrl+P to print</em></p>
            </div>
            
            <div class="header">
                <div class="logo">
                    <img src="{logo_src}" alt="LOGO" style="max-width: 100%; max-height: 100%;" onerror="this.style.display='none'; this.parentNode.innerHTML='ALKHIDMAT<br>AGHOSH<br>HALA';">
                </div>
                <div class="header-text">
                    <h2 style="white-space: nowrap;">ALKHIDMAT SCHOOL, MANNAN & QAZI CAMPUS, HALA</h2>
                    <p>Near Bhitshah Bus Stop, Main National Highway 05, Hala</p>
                    <p>Phone # 0333 112 0663    Email: alkhidmataghosh@alkhidmat.org</p>
                    <h3>Final Term Examination - March 2025</h3>
                    <p>Report Card for Academic Year 2024-25</p>
                </div>
            </div>
            
            <div class="student-info">
                <span>Student's Name: {student['name']}</span>
                <span>Class: {student['class']}</span>
                <span>Aghosh ID: {self.current_student_id}</span>
            </div>
            
            <table>
                <tr>
                    <th>Sr.#</th>
                    <th>Subjects</th>
                    <th>1st Term<br>100/50</th>
                    <th>Weighted<br>20%</th>
                    <th>2nd Term<br>100/50</th>
                    <th>Weighted<br>80%</th>
                    <th>Aggregate</th>
                    <th>Percentage</th>
                    <th>Grade</th>
                    <th>Remarks</th>
                </tr>
                {rows_html}
                <tr class="total-row">
                    <td></td>
                    <td>Total [{class_total_marks}]</td>
                    <td>{total_mid}</td>
                    <td></td>
                    <td>{total_final}</td>
                    <td></td>
                    <td>{total_agg:.0f}</td>
                    <td>{overall_perc:.0f}</td>
                    <td></td>
                    <td></td>
                </tr>
            </table>
            
            <div class="footer">
                <div class="grading">
                    <strong>Grading System</strong><br>
                    <table style="margin: 5px auto; text-align: center; font-size: 10px;">
                        <tr>
                            <td>90-99</td><td>80-89</td><td>70-79</td><td>60-69</td><td>50-59</td><td>39-49</td><td>0-39</td>
                        </tr>
                        <tr>
                            <td>A+</td><td>A</td><td>B</td><td>C</td><td>D</td><td>E</td><td>F</td>
                        </tr>
                    </table>
                </div>
                <div class="results">
                    <table style="margin: 0 auto; text-align: center; font-size: 11px; border: 1px solid black;">
                        <tr><td style="border: 1px solid black; padding: 3px;"><strong>Overall %:</strong> {overall_perc:.0f}%</td></tr>
                        <tr><td style="border: 1px solid black; padding: 3px;"><strong>Grade:</strong> {overall_grade}</td></tr>
                        <tr><td style="border: 1px solid black; padding: 3px;"><strong>Result:</strong> {pass_fail}</td></tr>
                    </table>
                </div>
            </div>
            
            <div class="teacher-remarks">
                <strong>Teacher's Remarks:</strong><br>
                {combined_remarks}
            </div>
            
            <div class="signatures">
                <div class="signature">
                    <div class="signature-line"></div>
                    Class Teacher's Sign
                </div>
                <div class="signature">
                    <div class="signature-line"></div>
                    Principal's Sign
                </div>
                <div class="signature">
                    <div class="signature-line"></div>
                    Guardian's Sign
                </div>
            </div>
        </body>
        </html>
        """
        return html_content
    
    def print_report(self):
        if not hasattr(self, 'current_student_id'):
            messagebox.showwarning("Warning", "Please select a student first")
            return
        
        try:
            student = self.students[self.current_student_id]
            subjects = self.get_subjects(student['class'], self.current_student_id)
            html_content = self.create_html_report(student, subjects)
            
            student_num = self.current_student_id.replace('AAH-', '').replace(' ', '')
            html_file = f"../Output_Files/report_card_{student_num}.html"
            
            with open(html_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            file_path = os.path.abspath(html_file)
            success = False
            
            browsers = ['microsoft-edge', 'google-chrome', 'firefox', 'chromium-browser']
            for browser in browsers:
                try:
                    subprocess.run([browser, file_path], check=True)
                    success = True
                    break
                except:
                    continue
            
            if not success:
                try:
                    subprocess.run(['xdg-open', file_path], check=True)
                    success = True
                except:
                    pass
            
            if success:
                messagebox.showinfo("Success", f"Report card opened!\nFile: {html_file}\nUse Ctrl+P to print")
            else:
                messagebox.showinfo("File Created", f"Report card saved as: {html_file}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not create report: {str(e)}")
        if not hasattr(self, 'current_student_id'):
            messagebox.showwarning("Warning", "Please select a student first")
            return
        
        try:
            student = self.students[self.current_student_id]
            subjects = self.get_subjects(student['class'])
            html_content = self.create_html_report(student, subjects)
            
            student_num = self.current_student_id.replace('AAH-', '').replace(' ', '')
            html_file = f"../Output_Files/report_card_{student_num}.html"
            
            with open(html_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            file_path = os.path.abspath(html_file)
            success = False
            
            browsers = ['microsoft-edge', 'google-chrome', 'firefox', 'chromium-browser']
            for browser in browsers:
                try:
                    subprocess.run([browser, file_path], check=True)
                    success = True
                    break
                except:
                    continue
            
            if not success:
                try:
                    subprocess.run(['xdg-open', file_path], check=True)
                    success = True
                except:
                    pass
            
            if success:
                messagebox.showinfo("Success", f"Report card opened!\nFile: {html_file}\nUse Ctrl+P to print")
            else:
                messagebox.showinfo("File Created", f"Report card saved as: {html_file}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not create report: {str(e)}")
        if not hasattr(self, 'current_student_id'):
            messagebox.showwarning("Warning", "Please select a student first")
            return
        
        try:
            student = self.students[self.current_student_id]
            subjects = self.get_subjects(student['class'])
            html_content = self.create_html_report(student, subjects)
            
            student_num = self.current_student_id.replace('AAH-', '').replace(' ', '')
            html_file = f"../Output_Files/report_card_{student_num}.html"
            
            with open(html_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            file_path = os.path.abspath(html_file)
            success = False
            
            browsers = ['microsoft-edge', 'google-chrome', 'firefox', 'chromium-browser']
            for browser in browsers:
                try:
                    subprocess.run([browser, file_path], check=True)
                    success = True
                    break
                except:
                    continue
            
            if not success:
                try:
                    subprocess.run(['xdg-open', file_path], check=True)
                    success = True
                except:
                    pass
            
            if success:
                messagebox.showinfo("Success", f"Report card opened!\nFile: {html_file}\nUse Ctrl+P to print")
            else:
                messagebox.showinfo("File Created", f"Report card saved as: {html_file}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not create report: {str(e)}")
    
    def create_report(self):
        # Header
        header_frame = tk.Frame(self.report_frame, bg='white', relief='solid', borderwidth=1)
        header_frame.pack(fill='x', padx=3, pady=3)
        
        top_row = tk.Frame(header_frame, bg='white')
        top_row.pack(fill='x', pady=8)
        
        try:
            logo_img = Image.open("../Assets/Aghos logo.png")
            logo_img = logo_img.resize((100, 100))
            self.logo_photo = ImageTk.PhotoImage(logo_img)
            logo_label = tk.Label(top_row, image=self.logo_photo, bg='white')
            logo_label.pack(side='left', padx=15)
        except:
            logo_frame = tk.Frame(top_row, bg='lightblue', width=100, height=100, relief='solid', borderwidth=1)
            logo_frame.pack(side='left', padx=15)
            logo_frame.pack_propagate(False)
            tk.Label(logo_frame, text="LOGO", bg='lightblue', font=('Arial', 12)).pack(expand=True)
        
        info_frame = tk.Frame(top_row, bg='white')
        info_frame.pack(side='left', fill='x', expand=True, padx=10)
        
        tk.Label(info_frame, text="ALKHIDMAT SCHOOL, MANNAN & QAZI CAMPUS, HALA", 
                font=('Arial', 16, 'bold'), bg='white', anchor='center').pack(pady=2, fill='x')
        tk.Label(info_frame, text="Near Bhitshah Bus Stop, Main National Highway 05, Hala", 
                font=('Arial', 10), bg='white', anchor='center').pack(fill='x')
        tk.Label(info_frame, text="Phone # 0333 112 0663    Email: alkhidmataghosh@alkhidmat.org", 
                font=('Arial', 10), bg='white', anchor='center').pack(fill='x')
        
        tk.Label(header_frame, text="Final Term Examination - March 2025", 
                font=('Arial', 14, 'bold'), bg='white', anchor='center').pack(pady=5, fill='x')
        tk.Label(header_frame, text="Report Card for Academic Year 2024-25", 
                font=('Arial', 12), bg='white', anchor='center').pack(pady=(0,8), fill='x')
        
        # Student info
        student_frame = tk.Frame(self.report_frame, bg='white', relief='solid', borderwidth=1)
        student_frame.pack(fill='x', padx=3, pady=3)
        
        info_grid = tk.Frame(student_frame, bg='white')
        info_grid.pack(fill='x', padx=15, pady=10)
        
        tk.Label(info_grid, text="Student's Name:", font=('Arial', 11, 'bold'), bg='white').grid(row=0, column=0, sticky='w')
        self.name_label = tk.Label(info_grid, text="", font=('Arial', 11, 'bold'), bg='white', width=22, anchor='w')
        self.name_label.grid(row=0, column=1, sticky='w', padx=5)
        
        tk.Label(info_grid, text="Class:", font=('Arial', 11, 'bold'), bg='white').grid(row=0, column=2, sticky='w', padx=(25,5))
        self.class_label = tk.Label(info_grid, text="", font=('Arial', 11, 'bold'), bg='white', width=5, anchor='center')
        self.class_label.grid(row=0, column=3, sticky='w', padx=5)
        
        tk.Label(info_grid, text="Aghosh ID:", font=('Arial', 11, 'bold'), bg='white').grid(row=0, column=4, sticky='w', padx=(25,5))
        self.id_label = tk.Label(info_grid, text="", font=('Arial', 11, 'bold'), bg='white', width=15, anchor='center')
        self.id_label.grid(row=0, column=5, sticky='w', padx=5)
        
        self.table_frame = tk.Frame(self.report_frame, bg='white')
        self.table_frame.pack(fill='x', padx=3, pady=5)
        
        self.create_footer()
    
    def create_footer(self):
        footer_frame = tk.Frame(self.report_frame, bg='white')
        footer_frame.pack(fill='x', padx=3, pady=10)
        
        # Grading system
        left_frame = tk.Frame(footer_frame, bg='white', relief='solid', borderwidth=1, width=380, height=80)
        left_frame.pack(side='left', fill='y', padx=5)
        left_frame.pack_propagate(False)
        
        tk.Label(left_frame, text="Grading System", font=('Arial', 10, 'bold'), bg='white').pack(pady=3)
        
        grade_line1 = "90-99  80-89  70-79  60-69  50-59  39-49  0-39"
        grade_line2 = " A+     A      B      C      D      E     F"
        
        tk.Label(left_frame, text=grade_line1, font=('Arial', 8), bg='white').pack()
        tk.Label(left_frame, text=grade_line2, font=('Arial', 8), bg='white').pack()
        
        # Right side
        right_frame = tk.Frame(footer_frame, bg='white', relief='solid', borderwidth=1)
        right_frame.pack(side='right', fill='both', expand=True, padx=5)
        
        pass_fail_row = tk.Frame(right_frame, bg='white')
        pass_fail_row.pack(fill='x', padx=8, pady=5)
        
        tk.Label(pass_fail_row, text="Remarks:", font=('Arial', 10, 'bold'), bg='white').pack(side='left')
        self.pass_fail_label = tk.Label(pass_fail_row, text="Pass", font=('Arial', 10, 'bold'), bg='white')
        self.pass_fail_label.pack(side='right')
        
        grade_row = tk.Frame(right_frame, bg='white')
        grade_row.pack(fill='x', padx=8, pady=2)
        
        tk.Label(grade_row, text="Grade:", font=('Arial', 10), bg='white').pack(side='left')
        self.grade_label = tk.Label(grade_row, text="A+", font=('Arial', 10), bg='white')
        self.grade_label.pack(side='right')
        
        percentage_row = tk.Frame(right_frame, bg='white')
        percentage_row.pack(fill='x', padx=8, pady=2)
        
        tk.Label(percentage_row, text="Overall %:", font=('Arial', 10), bg='white').pack(side='left')
        self.overall_label = tk.Label(percentage_row, text="", font=('Arial', 10), bg='white')
        self.overall_label.pack(side='right')
        
        # Teacher's Remarks
        remarks_frame = tk.Frame(self.report_frame, bg='white', relief='solid', borderwidth=1)
        remarks_frame.pack(fill='x', padx=3, pady=5)
        
        tk.Label(remarks_frame, text="Teacher's Remarks:", font=('Arial', 10, 'bold'), bg='white').pack(anchor='w', padx=8, pady=3)
        self.teacher_remarks = tk.Text(remarks_frame, height=3, font=('Arial', 9), bg='white')
        self.teacher_remarks.pack(fill='x', padx=8, pady=3)
        self.teacher_remarks.insert(1.0, "Student has shown good performance in academics.")
        
        tk.Label(self.report_frame, text="", bg='white', height=2).pack()
        
        # Signatures
        sig_frame = tk.Frame(self.report_frame, bg='white')
        sig_frame.pack(fill='x', padx=3, pady=10)
        
        sig_left = tk.Frame(sig_frame, bg='white')
        sig_left.pack(side='left', fill='x', expand=True)
        tk.Label(sig_left, text="_" * 20, font=('Arial', 10), bg='white').pack()
        tk.Label(sig_left, text="Class Teacher's Sign", font=('Arial', 10), bg='white').pack()
        
        sig_center = tk.Frame(sig_frame, bg='white')
        sig_center.pack(side='left', fill='x', expand=True)
        tk.Label(sig_center, text="_" * 20, font=('Arial', 10), bg='white').pack()
        tk.Label(sig_center, text="Principal's Sign", font=('Arial', 10), bg='white').pack()
        
        sig_right = tk.Frame(sig_frame, bg='white')
        sig_right.pack(side='left', fill='x', expand=True)
        tk.Label(sig_right, text="_" * 20, font=('Arial', 10), bg='white').pack()
        tk.Label(sig_right, text="Guardian's Sign", font=('Arial', 10), bg='white').pack()
    
    def create_html_report(self, student, subjects):
        total_mid = 0
        total_final = 0
        total_agg = 0
        calculation_subjects = 0
        rows_html = ""
        
        for i, subject in enumerate(subjects):
            if self.is_computer_subject(subject):
                # Computer subject - show grades but calculate differently
                mid_grade = self.get_computer_grade(student, i, 'mid')
                final_grade = self.get_computer_grade(student, i, 'final')
                
                # Convert grades to percentage for calculation
                mid_perc = self.grade_to_percentage(mid_grade)
                final_perc = self.grade_to_percentage(final_grade)
                
                # Calculate average of both grades
                avg_perc = (mid_perc + final_perc) / 2
                avg_grade = self.percentage_to_grade(avg_perc)
                
                remarks = self.get_remarks_for_grade(avg_grade)
                
                rows_html += f"""
                <tr>
                    <td>{i+1}</td>
                    <td>{subject}</td>
                    <td>{mid_grade}</td>
                    <td>-</td>
                    <td>{final_grade}</td>
                    <td>-</td>
                    <td>{avg_grade}</td>
                    <td>{avg_perc:.0f}</td>
                    <td>{avg_grade}</td>
                    <td>{remarks}</td>
                </tr>
                """
                
                total_mid += mid_perc
                total_final += final_perc
                total_agg += avg_perc
                calculation_subjects += 1
            else:
                mid = student['mid_marks'][i] if i < len(student['mid_marks']) else 0
                final = student['final_marks'][i] if i < len(student['final_marks']) else 0
                
                # Determine max marks for this subject based on class and subject
                if student['class'] in ['II', 'III']:
                    # English, Urdu, GK, Maths = 100 marks; Islamiat, Sindhi = 50 marks
                    max_marks = 100 if subject in ['English', 'Urdu', 'GK', 'Mathematics'] else 50
                elif student['class'] in ['IV', 'V', 'VI', 'VII', 'VIII']:
                    # English, Urdu, Science, Maths = 100 marks; S.S, Islamiat, Sindhi = 50 marks
                    max_marks = 100 if subject in ['English', 'Urdu', 'Science', 'Mathematics'] else 50
                elif student['class'] in ['IX', 'X']:
                    # English, Urdu, Maths, Bio, Phy, Che = 100 marks; Islamiat/PS = 50 marks
                    max_marks = 100 if subject in ['English', 'Urdu', 'Mathematics', 'Biology', 'Physics', 'Chemistry'] else 50
                else:
                    max_marks = 100
                
                w_mid = (mid * 20) / 100
                w_final = (final * 80) / 100
                agg = w_mid + w_final
                perc = (agg / max_marks) * 100  # Calculate percentage based on max marks
                
                grade = 'A+' if perc >= 91 else 'A' if perc >= 80 else 'B' if perc >= 70 else 'C' if perc >= 60 else 'D' if perc >= 50 else 'E' if perc >= 35 else 'F'
                remarks = self.get_remarks_for_grade(grade)
                
                rows_html += f"""
                <tr>
                    <td>{i+1}</td>
                    <td>{subject}</td>
                    <td>{mid}</td>
                    <td>{w_mid:.1f}</td>
                    <td>{final}</td>
                    <td>{w_final:.1f}</td>
                    <td>{agg:.0f}</td>
                    <td>{perc:.0f}</td>
                    <td>{grade}</td>
                    <td>{remarks}</td>
                </tr>
                """
                
                total_mid += mid
                total_final += final
                total_agg += agg
                calculation_subjects += 1
        
        overall_perc = total_agg/calculation_subjects if calculation_subjects > 0 else 0
        overall_grade = 'A+' if overall_perc >= 91 else 'A' if overall_perc >= 80 else 'B' if overall_perc >= 70 else 'C'
        pass_fail = "Pass" if overall_perc >= 40 else "Fail"
        
        # Calculate correct total marks dynamically
        class_total_marks = self.get_class_total_marks(student['class'], subjects)
        teacher_remarks = self.teacher_remarks.get(1.0, tk.END).strip() if hasattr(self, 'teacher_remarks') else "Student has shown good performance."
        
        # Add automatic performance remarks based on percentage
        performance_remark = self.get_remarks_by_percentage(overall_perc)
        combined_remarks = f"Performance: {performance_remark}. {teacher_remarks}"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Report Card - {self.current_student_id}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; border: 2px solid black; padding: 20px; margin-bottom: 20px; }}
                .student-info {{ border: 2px solid black; padding: 15px; margin-bottom: 20px; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                th, td {{ border: 1px solid black; padding: 8px; text-align: center; }}
                th {{ background-color: lightgray; font-weight: bold; }}
                .total-row {{ background-color: lightgray; font-weight: bold; }}
                .footer {{ display: flex; margin-top: 20px; }}
                .grading {{ border: 1px solid black; padding: 10px; width: 40%; }}
                .remarks {{ border: 1px solid black; padding: 10px; flex: 1; margin-left: 10px; }}
                .teacher-remarks {{ border: 1px solid black; padding: 10px; margin: 10px 0; }}
                .signatures {{ margin-top: 50px; display: flex; justify-content: space-between; }}
                .signature {{ text-align: center; }}
                @media print {{ body {{ margin: 0; }} }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>ALKHIDMAT SCHOOL, MANNAN & QAZI CAMPUS, HALA</h2>
                <p>Near Bhitshah Bus Stop, Main National Highway 05, Hala</p>
                <p>Phone # 0333 112 0663    Email: alkhidmataghosh@alkhidmat.org</p>
                <h3>Final Term Examination - March 2025</h3>
                <p>Report Card for Academic Year 2024-25</p>
            </div>
            
            <div class="student-info">
                <strong>Student's Name:</strong> {student['name']} &nbsp;&nbsp;&nbsp;
                <strong>Class:</strong> {student['class']} &nbsp;&nbsp;&nbsp;
                <strong>Aghosh ID:</strong> {self.current_student_id}
            </div>
            
            <table>
                <tr>
                    <th>Sr. #</th>
                    <th>Subjects</th>
                    <th>1st Term<br>100/50</th>
                    <th>Weighted<br>20%</th>
                    <th>2nd Term<br>100/50</th>
                    <th>Weighted<br>80%</th>
                    <th>Aggregate</th>
                    <th>Percentage</th>
                    <th>Grade</th>
                    <th>Remarks</th>
                </tr>
                {rows_html}
                <tr class="total-row">
                    <td></td>
                    <td>Total [{class_total_marks}]</td>
                    <td>{total_mid}</td>
                    <td></td>
                    <td>{total_final}</td>
                    <td></td>
                    <td>{total_agg:.0f}</td>
                    <td>{overall_perc:.0f}</td>
                    <td></td>
                    <td></td>
                </tr>
            </table>
            
            <div class="footer">
                <div class="grading">
                    <strong>Grading System</strong><br>
                    <table style="margin: 5px auto; text-align: center; font-size: 10px;">
                        <tr>
                            <td>90-99</td><td>80-89</td><td>70-79</td><td>60-69</td><td>50-59</td><td>39-49</td><td>0-39</td>
                        </tr>
                        <tr>
                            <td>A+</td><td>A</td><td>B</td><td>C</td><td>D</td><td>E</td><td>F</td>
                        </tr>
                    </table>
                </div>
                <div class="remarks">
                    <table style="margin: 0 auto; text-align: center; font-size: 11px; border: 1px solid black;">
                        <tr><td style="border: 1px solid black; padding: 3px;"><strong>Overall %:</strong> {overall_perc:.0f}%</td></tr>
                        <tr><td style="border: 1px solid black; padding: 3px;"><strong>Grade:</strong> {overall_grade}</td></tr>
                        <tr><td style="border: 1px solid black; padding: 3px;"><strong>Result:</strong> {pass_fail}</td></tr>
                    </table>
                </div>
            </div>
            
            <div class="teacher-remarks">
                <strong>Teacher's Remarks:</strong><br>
                {combined_remarks}
            </div>
            
            <div class="signatures">
                <div class="signature">
                    ____________________<br>
                    Class Teacher's Sign
                </div>
                <div class="signature">
                    ____________________<br>
                    Principal's Sign
                </div>
                <div class="signature">
                    ____________________<br>
                    Guardian's Sign
                </div>
            </div>
        </body>
        </html>
        """
        return html_content

if __name__ == "__main__":
    root = tk.Tk()
    
    # Ensure window appears in taskbar
    root.withdraw()  # Hide window temporarily
    root.deiconify()  # Show window again (forces taskbar appearance)
    
    app = ReportCardFixedTotals(root)
    root.mainloop()
