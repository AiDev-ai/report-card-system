import openpyxl

def search_006():
    try:
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
        final_wb = openpyxl.load_workbook("Excel_Data/Exams/Final Term.xlsx")
        
        class_sheets = ['Class  II', 'Class III', 'Class IV', 'Class V', 'Class VI', 'Class VII', 'Class VIII', 'Class IX', 'Class X']
        
        for sheet_name in class_sheets:
            if sheet_name in mid_wb.sheetnames:
                mid_sheet = mid_wb[sheet_name]
                final_sheet = final_wb[sheet_name]
                
                for row in range(1, 100):
                    for col in ['A', 'B', 'C']:
                        cell_val = mid_sheet[f'{col}{row}'].value
                        if cell_val and '006' in str(cell_val):
                            next_col = chr(ord(col) + 1)
                            name_val = mid_sheet[f'{next_col}{row}'].value
                            print(f"Found: {cell_val} - {name_val} in {sheet_name}")
                            print(f"Row: {row}, ID Column: {col}, Name Column: {next_col}")
                            
                            # Show all marks
                            print("Mid Term marks (D-J):")
                            for i, mark_col in enumerate(['D', 'E', 'F', 'G', 'H', 'I', 'J']):
                                mark_val = mid_sheet[f'{mark_col}{row}'].value
                                print(f"  Index {i} ({mark_col}): {mark_val} ({type(mark_val).__name__})")
                            
                            print("Final Term marks (D-J):")
                            for i, mark_col in enumerate(['D', 'E', 'F', 'G', 'H', 'I', 'J']):
                                mark_val = final_sheet[f'{mark_col}{row}'].value
                                print(f"  Index {i} ({mark_col}): {mark_val} ({type(mark_val).__name__})")
                            return
        
        print("Student with 006 not found")
        
    except Exception as e:
        print(f"Error: {e}")

search_006()
