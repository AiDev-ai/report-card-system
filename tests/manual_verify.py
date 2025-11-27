#!/usr/bin/env python3
"""
Manual Verification Tool - Spot Check Calculations
=================================================
Use this to manually verify any student's calculations
"""

import openpyxl
import sys

def manual_verify_student(student_id):
    """Manually verify a specific student's calculations"""
    
    print(f"🔍 Manual Verification for Student ID: {student_id}")
    print("=" * 60)
    
    # Load Excel files
    try:
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
        final_wb = openpyxl.load_workbook("Excel_Data/Exams/Final Term.xlsx")
    except Exception as e:
        print(f"❌ Error loading Excel files: {e}")
        return False
    
    # Find student in Excel sheets
    class_sheets = ['Class  II', 'Class III', 'Class IV', 'Class V', 'Class VI', 'Class VII', 'Class VIII', 'Class IX', 'Class X']
    
    student_found = False
    
    for sheet_name in class_sheets:
        if sheet_name in mid_wb.sheetnames and sheet_name in final_wb.sheetnames:
            mid_sheet = mid_wb[sheet_name]
            final_sheet = final_wb[sheet_name]
            
            for row in range(1, 200):
                # Check for student ID in columns A, B, C
                for col in ['A', 'B', 'C']:
                    cell_val = mid_sheet[f'{col}{row}'].value
                    if cell_val and str(cell_val).strip() == student_id:
                        student_found = True
                        
                        # Get student name
                        next_col = chr(ord(col) + 1)
                        student_name = mid_sheet[f'{next_col}{row}'].value
                        student_class = sheet_name.replace('Class ', '').strip()
                        
                        print(f"📋 Found: {student_name} (Class {student_class})")
                        print(f"📍 Location: {sheet_name}, Row {row}")
                        print()
                        
                        # Get subjects for this class
                        if student_class in ['II', 'III']:
                            subjects = ['English', 'Urdu', 'Mathematics', 'GK', 'Computer']
                            total_marks = 500  # Now includes Computer
                        elif student_class in ['IV', 'V', 'VI', 'VII', 'VIII']:
                            subjects = ['English', 'Urdu', 'Science', 'Mathematics', 'Computer', 'Islamiat', 'Social Studies']
                            total_marks = 700  # Now includes Computer
                        elif student_class in ['IX', 'X']:
                            subjects = ['English', 'Urdu', 'Physics', 'Mathematics', 'Computer', 'Islamiat']
                            total_marks = 600  # Now includes Computer
                        
                        print("📊 RAW DATA FROM EXCEL:")
                        print("-" * 40)
                        
                        # Extract marks from columns D through J
                        raw_mid_marks = []
                        raw_final_marks = []
                        
                        for i, mark_col in enumerate(['D', 'E', 'F', 'G', 'H', 'I', 'J']):
                            mid_val = mid_sheet[f'{mark_col}{row}'].value
                            final_val = final_sheet[f'{mark_col}{row}'].value
                            
                            subject_name = ""
                            if i < len(subjects):
                                subject_name = f" ({subjects[i]})"
                            elif mark_col == 'H' and student_class in ['II', 'III']:
                                subject_name = " (Computer)"
                            
                            print(f"Column {mark_col}: Mid={mid_val}, Final={final_val}{subject_name}")
                            
                            raw_mid_marks.append(mid_val)
                            raw_final_marks.append(final_val)
                        
                        print("\n🧮 STEP-BY-STEP CALCULATIONS:")
                        print("-" * 40)
                        
                        total_mid = 0
                        total_final = 0
                        total_agg = 0
                        calculation_subjects = 0
                        
                        for i, subject in enumerate(subjects):
                            if student_class in ['II', 'III'] and subject == 'Computer':
                                # For Class II and III, Computer is at column H (index 4)
                                computer_index = 4
                                mid = raw_mid_marks[computer_index] if isinstance(raw_mid_marks[computer_index], (int, float)) else 0
                                final = raw_final_marks[computer_index] if isinstance(raw_final_marks[computer_index], (int, float)) else 0
                            else:
                                mid = raw_mid_marks[i] if isinstance(raw_mid_marks[i], (int, float)) else 0
                                final = raw_final_marks[i] if isinstance(raw_final_marks[i], (int, float)) else 0
                            
                            if subject == 'Computer':
                                # Computer grade - show grades, calculate average
                                if student_class in ['II', 'III']:
                                    # Check column H for Class II and III
                                    mid_grade = raw_mid_marks[4] if isinstance(raw_mid_marks[4], str) and raw_mid_marks[4].strip() in ['A+', 'A', 'B', 'C', 'D', 'E', 'F'] else 'C'
                                    final_grade = raw_final_marks[4] if isinstance(raw_final_marks[4], str) and raw_final_marks[4].strip() in ['A+', 'A', 'B', 'C', 'D', 'E', 'F'] else 'C'
                                else:
                                    # For other classes, use the regular index
                                    mid_grade = raw_mid_marks[i] if isinstance(raw_mid_marks[i], str) and raw_mid_marks[i].strip() in ['A+', 'A', 'B', 'C', 'D', 'E', 'F'] else 'C'
                                    final_grade = raw_final_marks[i] if isinstance(raw_final_marks[i], str) and raw_final_marks[i].strip() in ['A+', 'A', 'B', 'C', 'D', 'E', 'F'] else 'C'
                                
                                # Convert grades to percentages for calculation
                                grade_to_perc = {'A+': 95, 'A': 85, 'B': 75, 'C': 65, 'D': 55, 'E': 45, 'F': 25}
                                mid_perc = grade_to_perc.get(mid_grade, 65)
                                final_perc = grade_to_perc.get(final_grade, 65)
                                
                                # Calculate average percentage
                                avg_perc = (mid_perc + final_perc) / 2
                                
                                # Convert back to grade
                                if avg_perc >= 91: avg_grade = 'A+'
                                elif avg_perc >= 80: avg_grade = 'A'
                                elif avg_perc >= 70: avg_grade = 'B'
                                elif avg_perc >= 60: avg_grade = 'C'
                                elif avg_perc >= 50: avg_grade = 'D'
                                elif avg_perc >= 35: avg_grade = 'E'
                                else: avg_grade = 'F'
                                
                                print(f"{i+1:2d}. {subject:15s} | Mid: {mid_grade:2s} | Final: {final_grade:2s}")
                                print(f"    Weighted Mid: - (grade display)")
                                print(f"    Weighted Final: - (grade display)")
                                print(f"    Aggregate: {avg_grade} (average of {mid_perc} and {final_perc})")
                                print(f"    Percentage: {avg_perc:.0f}")
                                print()
                                
                                total_mid += mid_perc
                                total_final += final_perc
                                total_agg += avg_perc
                                calculation_subjects += 1
                            else:
                                # Regular subject calculation
                                w_mid = (mid * 20) / 100
                                w_final = (final * 80) / 100
                                agg = w_mid + w_final
                                
                                if agg >= 91: grade = 'A+'
                                elif agg >= 80: grade = 'A'
                                elif agg >= 70: grade = 'B'
                                elif agg >= 60: grade = 'C'
                                elif agg >= 50: grade = 'D'
                                elif agg >= 35: grade = 'E'
                                else: grade = 'F'
                                
                                print(f"{i+1:2d}. {subject:15s} | Mid: {mid:3.0f} | Final: {final:3.0f}")
                                print(f"    Weighted Mid: {mid} × 0.20 = {w_mid:5.1f}")
                                print(f"    Weighted Final: {final} × 0.80 = {w_final:5.1f}")
                                print(f"    Aggregate: {w_mid:4.1f} + {w_final:4.1f} = {agg:5.1f}")
                                print(f"    Grade: {grade}")
                                print()
                                
                                total_mid += mid
                                total_final += final
                                total_agg += agg
                                calculation_subjects += 1
                        
                        # Overall calculations
                        overall_perc = total_agg / calculation_subjects if calculation_subjects > 0 else 0
                        
                        if overall_perc >= 91: overall_grade = 'A+'
                        elif overall_perc >= 80: overall_grade = 'A'
                        elif overall_perc >= 70: overall_grade = 'B'
                        elif overall_perc >= 60: overall_grade = 'C'
                        else: overall_grade = 'C'
                        
                        pass_fail = "Pass" if overall_perc >= 40 else "Fail"
                        
                        print("🎯 FINAL CALCULATIONS:")
                        print("-" * 40)
                        print(f"Total Mid-term Marks: {total_mid}")
                        print(f"Total Final-term Marks: {total_final}")
                        print(f"Total Aggregate: {total_agg:.1f}")
                        print(f"Subjects in Calculation: {calculation_subjects}")
                        print(f"Overall Percentage: {total_agg:.1f} ÷ {calculation_subjects} = {overall_perc:.1f}%")
                        print(f"Overall Grade: {overall_grade}")
                        print(f"Pass/Fail: {pass_fail}")
                        print(f"Class Total Marks: {total_marks}")
                        
                        print("\n✅ VERIFICATION COMPLETE!")
                        return True
                
                if student_found:
                    break
        
        if student_found:
            break
    
    if not student_found:
        print(f"❌ Student ID '{student_id}' not found in any class sheet")
        return False

def list_all_students():
    """List all available student IDs"""
    print("📋 Available Student IDs:")
    print("=" * 30)
    
    try:
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return
    
    class_sheets = ['Class  II', 'Class III', 'Class IV', 'Class V', 'Class VI', 'Class VII', 'Class VIII', 'Class IX', 'Class X']
    
    for sheet_name in class_sheets:
        if sheet_name in mid_wb.sheetnames:
            mid_sheet = mid_wb[sheet_name]
            print(f"\n📚 {sheet_name}:")
            
            for row in range(1, 200):
                for col in ['A', 'B', 'C']:
                    cell_val = mid_sheet[f'{col}{row}'].value
                    if cell_val and str(cell_val).startswith('AAH'):
                        next_col = chr(ord(col) + 1)
                        name_val = mid_sheet[f'{next_col}{row}'].value
                        if name_val and not str(name_val).startswith('AAH'):
                            print(f"  {str(cell_val).strip()} - {str(name_val).strip()}")
                        break

if __name__ == "__main__":
    if len(sys.argv) > 1:
        if sys.argv[1] == "list":
            list_all_students()
        else:
            manual_verify_student(sys.argv[1])
    else:
        print("Manual Verification Tool")
        print("=" * 30)
        print("Usage:")
        print("  python3 manual_verify.py <student_id>  - Verify specific student")
        print("  python3 manual_verify.py list         - List all students")
        print()
        print("Examples:")
        print("  python3 manual_verify.py 'AAH- 170'")
        print("  python3 manual_verify.py 'AAH-99'")
        print("  python3 manual_verify.py list")
