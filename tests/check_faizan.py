import openpyxl

def check_faizan_data():
    try:
        mid_wb = openpyxl.load_workbook("Excel_Data/Exams/Mid Term.xlsx")
        final_wb = openpyxl.load_workbook("Excel_Data/Exams/Final Term.xlsx")
        
        class_sheets = ['Class  II', 'Class III', 'Class IV', 'Class V', 'Class VI', 'Class VII', 'Class VIII', 'Class IX', 'Class X']
        
        for sheet_name in class_sheets:
            if sheet_name in mid_wb.sheetnames and sheet_name in final_wb.sheetnames:
                mid_sheet = mid_wb[sheet_name]
                final_sheet = final_wb[sheet_name]
                
                print(f"\n=== Checking {sheet_name} ===")
                
                for row in range(1, 50):
                    try:
                        for col in ['A', 'B', 'C']:
                            cell_val = mid_sheet[f'{col}{row}'].value
                            if cell_val and 'AAH-007' in str(cell_val):
                                print(f"Found Faizan at row {row}, column {col}")
                                
                                # Get name
                                next_col = chr(ord(col) + 1)
                                name_val = mid_sheet[f'{next_col}{row}'].value
                                print(f"Name: {name_val}")
                                
                                # Get all marks from columns D to J
                                print("Mid Term marks:")
                                for mark_col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
                                    mark_val = mid_sheet[f'{mark_col}{row}'].value
                                    print(f"  Column {mark_col}: {mark_val} (type: {type(mark_val)})")
                                
                                print("Final Term marks:")
                                for mark_col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
                                    mark_val = final_sheet[f'{mark_col}{row}'].value
                                    print(f"  Column {mark_col}: {mark_val} (type: {type(mark_val)})")
                                
                                return
                    except:
                        continue
        
        print("Faizan AAH-007 not found in any sheet")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    check_faizan_data()
